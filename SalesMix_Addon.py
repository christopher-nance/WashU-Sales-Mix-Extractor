# Built by Christopher Nance for WashU Car Wash
# Version 5.0
# Sales Mix Report Generator

# Dependencies:
# > Python 3.10
# > Datetime
# > Calendar
# > re
# > copy
# > Pandas
# > Openpyxl
# > DRB Systems' SiteWatch 27


import pandas as pd
import openpyxl
from openpyxl.chart.plotarea import DataTable
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference, PieChart, ProjectedPieChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.chart.label import DataLabelList
from copy import copy
from datetime import datetime, date, timedelta
import re
import calendar


Corporation_Totals_Name = 'Corporation Totals'
global monthsInReport
monthsInReport = 1

ARM_Sold_Names = {
    "Express Wash": {
        "New Mnthly Express",
        "Mnth Exp 9.95 (BOGO)", # Added 10/16/23
        "New Mnthly Exp 9.95" # Added 10/16/23
    },
    "Clean Wash": {
        "Mnth Cln 9.95 (BOGO)",
        "Monthly Cln '21 9.95",
        "Monthly Cln '21 Sld",
    },
    "Protect Wash": {
        "Mnthly Protect Promo",
        "New Mnthly Protect",
    },
    "UShine Wash": {
        "Mnth USh 9.95 (BOGO)",
        "Monthly UShine Promo",
        "Monthly UShine Sld",
        "Mnth USh 9.95 (BOGO)" # Added 10/16/23
    },
}

ARM_Recharge_Names = {
    "Express Wash": {
        "New Mnthly Exp Rchg",
        "New Monthly Expr Rcg",
    },
    "Clean Wash": {
        "Monthly Cln '21 Rchg",
    },
    "Protect Wash": {
        "New Mnthly Prot Rchg",
        "Mnthly Prot 6mo Rchg",
        
    },
    "UShine Wash": {
        "Monthly UShine Rchg"
    },
}

ARM_Termination_Names = {
    "Express Wash": {
        "New Mnthly Exp NoRfn",
        "New Mnthly Exp Rfnd",
        "New Monthly Exp Rfnd",
        "New Monthly ExpNoRfn",
    },
    "Clean Wash": {
        "Monthly Cln '21 Rfnd",
        "MonthlyCln'21 NoRfnd",
        "New MonthlyCleNoRfnd"
        "New Monthly Cle Rfnd"
    },
    #"Shine Wash": {
    # "New Mnthly Shn NoRfn",
    # "New Mnthly Shn Rfnd"
    #},
    "Protect Wash": {
        "Mnthly Prot 6m NRfnd",
        "Mnthly Prot 6mo Rfnd",
        "New Mnthly Pro NoRfn",
        "New Mnthly Prot Rfnd",
    },
    "UShine Wash": {
        "Monthly UShine NRfnd",
        "Monthly UShine Rfnd",
    },
}

ARM_PKG_Names = {
    # Databook uses a modfied version which excludes the CLUB plans. 
    "Express Wash": {
        "New Mnthly Exp Rdmd",
        "New Monthly Expr Rdm",
        "3 Mo Express Rdmd",
        "CLUB-express Rdmd",
        "COMP-CLUB-xprs Rdmd",
    },
    "Clean Wash": {
        "3 Mo Clean Rdmd",
        "CLUB-clean Rdmd",
        "Monthly Cln '21 Rdmd",
        "COMP-CLUB-clean Rdmd",
        "W-Mon Clean V Rdmd",
    },
    "Protect Wash": {
        "3 Mo Protect Rdmd",
        "City Fire Club Rdmd",
        "CLUB-protect Rdmd",
        "COMP-CLUB-prot Rdmd",
        "Mnthly Prot 6mo Rdmd",
        "New Mnthy Prot Rdmd",
        "New Mnthly Prot Rdmd",
    },
    "UShine Wash": {
        "Monthly UShine Rdmd",
        "3 Mo UShine Rdmd",
        "CLUB-ushine Rdmd",
        "COMP-CLUB-ushin Rdmd",
        "Free Week UShine Rdm",
    },
}
Website_PKG_Names = {
    # The packages sold on the website differ in name and category than the other standard location packages. 
    # For instance, all website items (including ARM Sellers & Giftcard Sellers) are lumped into one "Website Sold" category
    "Retail": {
        "Express Wash": {
            
        },
        "Clean Wash": {
            "W-1-clean wash",
        },
        "Protect Wash": {
            "W-1-protect wash",
        },
        "UShine Wash": {
            "W-1-UShine wash",
        },
    },

    "Monthly": {
        "Express Wash": {
            "W-New MonthlyExprSld",
        },
        "Clean Wash": {
            "W-Mnthly Cln '21 Sld",
            "W-Mon Clean V Sld"
        },
        "Protect Wash": {
            "W-Unl. Prot 9.95 Sld",
            "W-Unl. protect Sld",
        },
        "UShine Wash": {
            "W-MonthlyUShine Sld",
            "W-UShine 9.95 Sld",
            "W-Mon UShine V Sld",
            "W-MonthlyUShine Sld"
        },
    }
}
blacklisted_items = [
    # Copied directly from the GSR CSV, hence the duplicate values.
    "Deposits",
    "XPT Cash Add/Remove",
    "XPT Cash Add/Remove",
    "XPT Cash Add/Remove",
    "XPT Cash Add/Remove",
    "XPT Cash Add/Remove",
    "House Acct Related",
    "House Acct Related",
    "House Acct Related",
    "House Acct Related",
    "House Acct Related",
    "House Acct Related",
    "Employee Accounts",
    "Employee Accounts",
    "Cash",
    "XPT Cash Over/Short",
    "XPT Chg Over/Short",
    "Credit Card",
    "Credit Card",
    "Credit Card",
    "Other Tenders",
    "Other Tenders",
    "House Accounts",
    "XPT Balancing",
    "Picture Mismatch",
    "Employees"
    #"Website Sold"
    "Terminate ARM Plan",
    "Terminate ARMNoRfnd"
]
blacklisted_sites = [
    # Add sites to this list to EXCLUDE them from the report.
    "Hub Office",
    #"wash*u - Centennial",
    "Zwash*u - Centennial Old",
    #"wash*u - Berwyn",
    #"wash*u - Joliet",
    #"wash*u - Des Plaines",
    #"wash*u - Des Plaine"
    #"Query Server"
]
    
Monthly_Total_Categories = [
    # Member revenue is calculated by adding amounts from these categories.
    "ARM Plans Sold", 
    "ARM Plans Recharged", 
    "Club Plans Sold", 
    "ARM Plans Terminated"
]
ARM_Member_Categories = [
    "ARM Plans Sold", 
    "ARM Plans Recharged", 
    "ARM Plans Terminated"
]
Discount_Categories = [
    "Website Discounts", 
    "Wash Discounts", 
    "Wash LPM Discounts",
    "Prepaid Redeemed"
]



def createSalesMixSheetWithVariance(gsrFilePath, templateFilePath, fileNameForParser, fileNameForParser2=None, historicalGSRFilePath=None, trendsFilePath=None, excludedLocations=None):
    #--> Create Dictionaries for looking up item names.
    # TODO: This needs to be attached to a database and allow insertions through the manager portal or a form.
    MONTHLY_STATS = {}
    COMBINED_STATS = {}
    COMBINED_SALES = {}
    MONTHLY_SALES = {}

    HIST_MONTHLY_STATS = {}
    HIST_COMBINED_STATS = {}
    HIST_COMBINED_SALES = {}
    HIST_MONTHLY_SALES = {}

    columnsToCopy = 11
    rowsToCopy = 32

    #--> Utility Functions
    def find_parent(json_obj, target_str, start_parent=None, current_parent=None):
        """
        Searches for a string in a nested JSON object (Python dict) starting from a specified parent,
        and returns the parent key under which the string resides.
        
        Parameters:
        - json_obj (dict): The JSON object to search through.
        - target_str (str): The string to look for.
        - start_parent (str): The parent key from where to start the search.
        - current_parent: The current parent key, used for recursion.
        
        Returns:
        - The parent key under which the string resides, or None if the string is not found.
        """
        if start_parent and current_parent is None:
            json_obj = json_obj.get(start_parent, {})
            
        for key, value in json_obj.items():
            if key == target_str:
                return current_parent
            if isinstance(value, dict):
                result = find_parent(value, target_str, None, key)
                if result:
                    return result
            elif isinstance(value, (list, set)):
                if target_str in value:
                    return key
        return None

    def move_worksheet_to_position(workbook, worksheet, position):
        """
        Move a worksheet to a specific position within a workbook.

        Parameters:
        - workbook: The Workbook object where the worksheet resides.
        - worksheet: The Worksheet object to be moved.
        - position: The position (0-based index) where the worksheet should be moved to.

        Returns:
        - True if successful, False otherwise.
        """
        if worksheet not in workbook._sheets:
            print("The worksheet is not in the given workbook.")
            return False

        if position < 0 or position >= len(workbook._sheets):
            print("Invalid position.")
            return False

        # Remove the original worksheet from the list
        workbook._sheets.remove(worksheet)

        # Insert it back at the specified position
        workbook._sheets.insert(position, worksheet)

        return True
    
    #--> Utility Functions (Graphs)
    def createMiniBarGraph(worksheet, START_COL, START_ROW):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Set the data range for the chart
        # Assuming data for the chart is in cells A4:B7
        values = Reference(ws, min_col=START_COL+1, min_row=START_ROW+3, max_col=START_COL+1, max_row=START_ROW+6)
        labels = Reference(ws, min_col=START_COL, min_row=START_ROW+3, max_row=START_ROW+6)
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(labels)

        # Set the dimensions of the chart
        chart.height = 4.05  # height
        chart.width = 11.35  # width

        # Remove the legend
        chart.legend = None

        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

        # Change bar colors (optional)
        colors = ['4682B4']  # Blue
        for i, s in enumerate(chart.series):
            s.graphicalProperties.solidFill = colors[i % len(colors)]
        
        CHART_PLACEMENT_COL = START_COL + 6
        CHART_PLACEMENT_ROW = START_ROW + 2

        # Add the chart to the worksheet
        ws.add_chart(chart, f"{get_column_letter(CHART_PLACEMENT_COL)}{CHART_PLACEMENT_ROW}")
    
    def createPieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None):
        ws = worksheet

        # Create a bar chart
        chart = PieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 9.5 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showLeaderLines = True


        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    def createCorporatePieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, showPercents=False):
        ws = worksheet

        # Create a bar chart
        chart = PieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 20
        chart.width = chartWidth if chartWidth != None else 20 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showLeaderLines = True
        if showPercents == True:
            chart.dLbls.showPercent = True


        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle

    def createBarGraph(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 19 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showLeaderLines = True
        chart.legend = None
        chart.y_axis.title = y_axisTitle if y_axisTitle != None else ''
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle

    def createCorporateBarGraph(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None, useDataTitle=True):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=useDataTitle)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 20
        chart.width = chartWidth if chartWidth != None else 20


        # Add data labels
        chart.dLbls = DataLabelList()
        #chart.dLbls.showVal = True
        #chart.dLbls.showLeaderLines = True
        #chart.legend = None
        chart.y_axis.title = y_axisTitle if y_axisTitle != None else ''

        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    def createPieInPieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None):
        ws = worksheet

        # Create a bar chart
        chart = ProjectedPieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 19

        chart.splitPos=4


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showLeaderLines = True
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    




    #########################################################################################################################################################################
    ############################################################################## GSR PARSING ##############################################################################
    #########################################################################################################################################################################
    #--> Gather Start/End Dates for the report

    # Read the CSV file to extract the start date and end date
    df = pd.read_csv(gsrFilePath, nrows=2)  # Reading only the first 2 rows

    def extract_dates_from_filepath(hist=False):
        global monthsInReport
        if hist == False:
            # Extract the file name from the file path
            filename = fileNameForParser
            
            # Use regular expression to find dates in the filename
            match = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})", filename)
            
            if match:
                start_date_str, end_date_str = match.groups()
                
                # Convert the extracted date strings to datetime objects
                start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d')

                start_date = start_date_dt.date()
                end_date = end_date_dt.date()

                # Convert the start and end dates from strings to datetime objects
                # Calculate the difference in days between the two dates
                delta = (end_date_dt- start_date_dt).days

                # Calculate the number of months
                monthsInReport = (delta / 30.44) if (delta / 30.44) > 1 else 1  # On average, a month is about 30.44 days, always make sure it is greater than 1 or else there will be in correct stats

                return start_date, end_date
            else:
                return None, None
        else:
            # Extract the file name from the file path
            filename = fileNameForParser2
            # Use regular expression to find dates in the filename
            match = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})", filename)
            
            if match:
                start_date_str, end_date_str = match.groups()
                
                # Convert the extracted date strings to datetime objects
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                return start_date, end_date
            else:
                return None, None

    start_date_hist, end_date_hist = extract_dates_from_filepath(hist=True)
    complete_date_str_hist = str(start_date_hist) + ' - ' + str(end_date_hist)

    start_date, end_date = extract_dates_from_filepath()
    short_date_str, complete_date_str = str(start_date) + ' - ' + str(end_date), str(start_date) + ' - ' + str(end_date)
        
    #--> Prepare dictionaries for data entry
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        print("Calculating for", site)
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']

        if site not in blacklisted_sites:
            if site not in MONTHLY_STATS:
                MONTHLY_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if site not in COMBINED_STATS:
                COMBINED_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            if Corporation_Totals_Name not in MONTHLY_STATS:
                MONTHLY_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if Corporation_Totals_Name not in COMBINED_STATS:
                COMBINED_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            
            if site not in MONTHLY_SALES:
                MONTHLY_SALES[site] = {"NET Sales": 0}
            if site not in COMBINED_SALES:
                COMBINED_SALES[site] = {"Discounts": 0, "NET Sales": 0}
            if Corporation_Totals_Name not in MONTHLY_SALES:
                MONTHLY_SALES[Corporation_Totals_Name] = {"NET Sales": 0}
            if Corporation_Totals_Name not in COMBINED_SALES:
                COMBINED_SALES[Corporation_Totals_Name] = {"Discounts": 0, "NET Sales": 0}

    #--> Gather TOTAL sales & discounts details
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            COMBINED_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
            COMBINED_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            if category in Discount_Categories:
                COMBINED_SALES[site]['Discounts'] += amount if pd.isna(amount) != True else 0
                COMBINED_SALES[Corporation_Totals_Name]['Discounts'] += amount if pd.isna(amount) != True else 0

            # Handle the Query Server
            if site == 'Query Server':
                if find_parent(Website_PKG_Names, item_name, start_parent="Monthly") != None or category in Monthly_Total_Categories:
                    MONTHLY_SALES['Query Server']['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            # Any other locations
            else: 
                if category in Monthly_Total_Categories:
                    MONTHLY_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0

            

    #--> Gather Sales Mix
    """
    Calculating Combined, Monthly, and Retail Stats:

    Components:
    - ARM Redeemer: A package leader item triggered when a FastPass Plan customer pays.
    - ARM Plan Packages: Package leader items contained within the ARM Redeemer.
    - Wash Package: Actual wash service (e.g., 'Express', 'Protect').
    - ARM Redeemer Discount: Makes the Wash Package cost $0 for FastPass Plan customers.

    Transaction Flow:
    1. FastPass Plan customer arrives, triggering the ARM Redeemer.
    2. The ARM Redeemer contains ARM Plan Packages, each of which includes:
        1. A Wash Package (e.g., 'Express', 'Protect').
        2. An ARM Redeemer Discount, setting the Wash Package cost to $0.

    Stat Calculations:
    - Combined Stats (Retail + Monthly): Both retail and monthly transactions are recorded under 'Basic Washes' because they both trigger a Wash Package.
    - Monthly Stats: Tracked when an ARM Redeemer is activated (exclusive to monthly FastPass Plan customers).
    - Retail Stats: Obtained by subtracting Monthly Stats from Combined Stats.

    Formula:
    Retail Stats = Combined Stats - Monthly Stats

    Template:
    The template is Excel based and therefore can be set up early to calculate the retail data based on the input from this script for combined and monthly.
    """
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    MonthlyAmounts = {}
    CombinedAmounts = {}
    AccurateMemberCountAdjuster = {}
    ChurnDataDictionary = {}
    ChurnTotal  = {}
    price_sum_dict = {}  # This will hold the sum of prices for each item for each site
    price_count_dict = {}  # This will hold the count of entries for each item for each site
    totalCorporateDiscont = 0

    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        count = row['Count']
        price = row['Price']
        quantity = row['Quantity']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            if ChurnDataDictionary.get(site) == None:
                ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            if ChurnDataDictionary.get(Corporation_Totals_Name) == None:
                ChurnDataDictionary[Corporation_Totals_Name] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            
            if AccurateMemberCountAdjuster.get(site) == None:
                AccurateMemberCountAdjuster[site] = 0
            if item_name in ['WEB Discontinue ARM', 'Discontinue ARM Plan']:
                if AccurateMemberCountAdjuster.get(site) != None:
                    AccurateMemberCountAdjuster[site] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    totalCorporateDiscont += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    ChurnDataDictionary[site]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    ChurnDataDictionary[Corporation_Totals_Name]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Recharge_Names, item_name) # Member Count = ARM Recharges
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(Website_PKG_Names, item_name, "Monthly") # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld | Gross New Members = ARM Sold
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[site][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                
            washPkg = find_parent(ARM_Termination_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld - ARM Terminations
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[site][washPkg]['Estimated Member Count'])
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'])
                ChurnDataDictionary[site]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
            
            if item_name == "Switch ARM Plan":
                ChurnDataDictionary[site]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            # Handle the Query Server
            if site == 'Query Server':
                # We're only going to show the amount of passes sold rather than the redemption for the
                # Query Server because you cannot redeem passes online.
                washPkg = find_parent(Website_PKG_Names, item_name, "Monthly")
                if washPkg != None:
                    MONTHLY_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members
    
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0)
                
                washPkg = find_parent(Website_PKG_Names, item_name, "Retail")
                if washPkg != None:
                    COMBINED_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        CombinedAmounts[washPkg] = []
                    CombinedAmounts[washPkg].append(price if pd.isna(price) != True else 0)

            # Any other locations
            else: 
                washPkg = find_parent(ARM_PKG_Names, item_name)
                # Calculate MONTHLY Stats using the redemption items
                if category in ["ARM Plans Redeemed", "Club Plans Redeemed"] and washPkg is not None:
                    MONTHLY_STATS[site][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[site][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[site][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS[site][washPkg]['Price'] += abs(price) if pd.isna(price) != True else 0

                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0) 
                elif category in ['Basic Washes'] and item_name is not None and item_name in COMBINED_STATS[site]:
                    COMBINED_STATS[site][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[site][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[site][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS[site][item_name]['Price'] += amount/quantity if pd.isna(quantity) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in CombinedAmounts:
                        CombinedAmounts[item_name] = []
                    CombinedAmounts[item_name].append(price if pd.isna(price) != True else 0)
                    '''if not site in price_sum_dict:
                        price_sum_dict[site] = {}
                    if not item_name in price_sum_dict[site]:
                        price_sum_dict[site][item_name] = []
                    price_sum_dict[site][item_name].append(price if pd.isna(price) != True else 0)'''

    for washPkg in ARM_Sold_Names:
        COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount']/COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount']/MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        #COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(CombinedAmounts[washPkg])/len(CombinedAmounts[washPkg])
        #MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(MonthlyAmounts[washPkg])/len(MonthlyAmounts[washPkg])
    AccurateMemberCountAdjuster[Corporation_Totals_Name] = totalCorporateDiscont

    for location, stats in ChurnDataDictionary.items():
        # ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
        ChurnTotal[location] = (ChurnDataDictionary[location]["Terminated"]+ChurnDataDictionary[location]["Discontinued"]) / (ChurnDataDictionary[location]["Recharged"]+ChurnDataDictionary[location]["Sold"]+ChurnDataDictionary[location]["Discontinued"]-ChurnDataDictionary[location]["Switched"])

    #########################################################################################################################################################################
    ####################################################################### HISTORICAL GSR PARSING ##########################################################################
    #########################################################################################################################################################################
    #--> Gather Start/End Dates for the report
        
    #--> Prepare dictionaries for data entry
    df = pd.read_csv(historicalGSRFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        print("Calculating for", site)
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']

        if site not in blacklisted_sites:
            if site not in HIST_MONTHLY_STATS:
                HIST_MONTHLY_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if site not in HIST_COMBINED_STATS:
                HIST_COMBINED_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            if Corporation_Totals_Name not in HIST_MONTHLY_STATS:
                HIST_MONTHLY_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if Corporation_Totals_Name not in HIST_COMBINED_STATS:
                HIST_COMBINED_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            
            if site not in HIST_MONTHLY_SALES:
                HIST_MONTHLY_SALES[site] = {"NET Sales": 0}
            if site not in HIST_COMBINED_SALES:
                HIST_COMBINED_SALES[site] = {"Discounts": 0, "NET Sales": 0}
            if Corporation_Totals_Name not in HIST_MONTHLY_SALES:
                HIST_MONTHLY_SALES[Corporation_Totals_Name] = {"NET Sales": 0}
            if Corporation_Totals_Name not in HIST_COMBINED_SALES:
                HIST_COMBINED_SALES[Corporation_Totals_Name] = {"Discounts": 0, "NET Sales": 0}

    #--> Gather TOTAL sales & discounts details
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(historicalGSRFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            HIST_COMBINED_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
            HIST_COMBINED_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            if category in Discount_Categories:
                HIST_COMBINED_SALES[site]['Discounts'] += amount if pd.isna(amount) != True else 0
                HIST_COMBINED_SALES[Corporation_Totals_Name]['Discounts'] += amount if pd.isna(amount) != True else 0

            # Handle the Query Server
            if site == 'Query Server':
                if find_parent(Website_PKG_Names, item_name, start_parent="Monthly") != None or category in Monthly_Total_Categories:
                    HIST_MONTHLY_SALES['Query Server']['NET Sales'] += amount if pd.isna(amount) != True else 0
                    HIST_MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            # Any other locations
            else: 
                if category in Monthly_Total_Categories:
                    HIST_MONTHLY_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
                    HIST_MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0

            

    #--> Gather Sales Mix
    """
    Calculating Combined, Monthly, and Retail Stats:

    Components:
    - ARM Redeemer: A package leader item triggered when a FastPass Plan customer pays.
    - ARM Plan Packages: Package leader items contained within the ARM Redeemer.
    - Wash Package: Actual wash service (e.g., 'Express', 'Protect').
    - ARM Redeemer Discount: Makes the Wash Package cost $0 for FastPass Plan customers.

    Transaction Flow:
    1. FastPass Plan customer arrives, triggering the ARM Redeemer.
    2. The ARM Redeemer contains ARM Plan Packages, each of which includes:
        1. A Wash Package (e.g., 'Express', 'Protect').
        2. An ARM Redeemer Discount, setting the Wash Package cost to $0.

    Stat Calculations:
    - Combined Stats (Retail + Monthly): Both retail and monthly transactions are recorded under 'Basic Washes' because they both trigger a Wash Package.
    - Monthly Stats: Tracked when an ARM Redeemer is activated (exclusive to monthly FastPass Plan customers).
    - Retail Stats: Obtained by subtracting Monthly Stats from Combined Stats.

    Formula:
    Retail Stats = Combined Stats - Monthly Stats

    Template:
    The template is Excel based and therefore can be set up early to calculate the retail data based on the input from this script for combined and monthly.
    """
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(historicalGSRFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    HIST_MonthlyAmounts = {}
    HIST_CombinedAmounts = {}
    HIST_AccurateMemberCountAdjuster = {}
    HIST_ChurnDataDictionary = {}
    HIST_ChurnTotal  = {}
    HIST_price_sum_dict = {}  # This will hold the sum of prices for each item for each site
    HIST_price_count_dict = {}  # This will hold the count of entries for each item for each site
    HIST_totalCorporateDiscont = 0

    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        count = row['Count']
        price = row['Price']
        quantity = row['Quantity']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            if HIST_ChurnDataDictionary.get(site) == None:
                HIST_ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            if HIST_ChurnDataDictionary.get(Corporation_Totals_Name) == None:
                HIST_ChurnDataDictionary[Corporation_Totals_Name] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            
            if HIST_AccurateMemberCountAdjuster.get(site) == None:
                HIST_AccurateMemberCountAdjuster[site] = 0
            if item_name in ['WEB Discontinue ARM', 'Discontinue ARM Plan']:
                if HIST_AccurateMemberCountAdjuster.get(site) != None:
                    HIST_AccurateMemberCountAdjuster[site] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    HIST_totalCorporateDiscont += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    HIST_ChurnDataDictionary[site]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    HIST_ChurnDataDictionary[Corporation_Totals_Name]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Recharge_Names, item_name) # Member Count = ARM Recharges
            if washPkg != None:
                HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_ChurnDataDictionary[site]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                HIST_ChurnDataDictionary[Corporation_Totals_Name]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(Website_PKG_Names, item_name, "Monthly") # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                HIST_ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld | Gross New Members = ARM Sold
            if washPkg != None:
                HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_MONTHLY_STATS[site][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members

                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                HIST_ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                
            washPkg = find_parent(ARM_Termination_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld - ARM Terminations
            if washPkg != None:
                HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'] = abs(HIST_MONTHLY_STATS[site][washPkg]['Estimated Member Count'])
                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] = abs(HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'])
                HIST_ChurnDataDictionary[site]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                HIST_ChurnDataDictionary[Corporation_Totals_Name]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
            
            if item_name == "Switch ARM Plan":
                HIST_ChurnDataDictionary[site]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                HIST_ChurnDataDictionary[Corporation_Totals_Name]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            # Handle the Query Server
            if site == 'Query Server':
                # We're only going to show the amount of passes sold rather than the redemption for the
                # Query Server because you cannot redeem passes online.
                washPkg = find_parent(Website_PKG_Names, item_name, "Monthly")
                if washPkg != None:
                    HIST_MONTHLY_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_MONTHLY_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_MONTHLY_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    HIST_MONTHLY_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0
                    HIST_MONTHLY_STATS['Query Server'][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members
    
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in HIST_MonthlyAmounts:
                        HIST_MonthlyAmounts[washPkg] = []
                    HIST_MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0)
                
                washPkg = find_parent(Website_PKG_Names, item_name, "Retail")
                if washPkg != None:
                    HIST_COMBINED_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_COMBINED_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_COMBINED_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    HIST_COMBINED_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0

                    HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in HIST_MonthlyAmounts:
                        HIST_CombinedAmounts[washPkg] = []
                    HIST_CombinedAmounts[washPkg].append(price if pd.isna(price) != True else 0)

            # Any other locations
            else: 
                washPkg = find_parent(ARM_PKG_Names, item_name)
                # Calculate MONTHLY Stats using the redemption items
                if category in ["ARM Plans Redeemed", "Club Plans Redeemed"] and washPkg is not None:
                    HIST_MONTHLY_STATS[site][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_MONTHLY_STATS[site][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_MONTHLY_STATS[site][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    HIST_MONTHLY_STATS[site][washPkg]['Price'] += abs(price) if pd.isna(price) != True else 0

                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in HIST_MonthlyAmounts:
                        HIST_MonthlyAmounts[washPkg] = []
                    HIST_MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0) 
                elif category in ['Basic Washes'] and item_name is not None and item_name in HIST_COMBINED_STATS[site]:
                    HIST_COMBINED_STATS[site][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_COMBINED_STATS[site][item_name]['Count'] += count if pd.isna(count) != True else 0
                    HIST_COMBINED_STATS[site][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    HIST_COMBINED_STATS[site][item_name]['Price'] += amount/quantity if pd.isna(quantity) != True else 0

                    HIST_COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    HIST_COMBINED_STATS[Corporation_Totals_Name][item_name]['Count'] += count if pd.isna(count) != True else 0
                    HIST_COMBINED_STATS[Corporation_Totals_Name][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in HIST_CombinedAmounts:
                        HIST_CombinedAmounts[item_name] = []
                    HIST_CombinedAmounts[item_name].append(price if pd.isna(price) != True else 0)
                    '''if not site in HIST_price_sum_dict:
                        HIST_price_sum_dict[site] = {}
                    if not item_name in HIST_price_sum_dict[site]:
                        HIST_price_sum_dict[site][item_name] = []
                    HIST_price_sum_dict[site][item_name].append(price if pd.isna(price) != True else 0)'''

    for washPkg in ARM_Sold_Names:
        HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount']/HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount']/HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        #HIST_COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(HIST_CombinedAmounts[washPkg])/len(HIST_CombinedAmounts[washPkg])
        #HIST_MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(HIST_MonthlyAmounts[washPkg])/len(HIST_MonthlyAmounts[washPkg])
    HIST_AccurateMemberCountAdjuster[Corporation_Totals_Name] = HIST_totalCorporateDiscont

    for location, stats in HIST_ChurnDataDictionary.items():
        # HIST_ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
        HIST_ChurnTotal[location] = (HIST_ChurnDataDictionary[location]["Terminated"]+HIST_ChurnDataDictionary[location]["Discontinued"]) / (HIST_ChurnDataDictionary[location]["Recharged"]+HIST_ChurnDataDictionary[location]["Sold"]+HIST_ChurnDataDictionary[location]["Discontinued"]-HIST_ChurnDataDictionary[location]["Switched"])

    #########################################################################################################################################################################
    ############################################################################## T&C PARSING ##############################################################################
    #########################################################################################################################################################################
    '''
        This can be used to extract the average churn over the period for a trends and comparisons report. The stats from this were off compared to SUDS so it was decided that we would use the 
        SUDS Formula to determine the Churn Rate using the items from the GSR alone. This eliminates the need for the Trends and Comparisons report.
        Formula we will use: Churn Rate = (Terminated + Discontinued) / (Recharged + Sold + Discontinued - Switched)


        ChurnByDay = {}
        ChurnTotal = {}
        if trendsFilePath is not None:
            df = pd.read_csv(trendsFilePath)
            for _, row in df.iterrows():
                date = row['Date']
                plan = row['Plan']
                site = row['Site']
                washes_per_day = row['Washes per day']
                upsells_dollars = row['Upsells (Dollars)']
                upsells = row['Upsells']
                plans_sold_dollars = row['Plans Sold (Dollars)']
                plans_sold = row['Plans Sold']
                recharges_dollars = row['Recharges (Dollars)']
                recharges = row['Recharges']
                plans_transferred_in = row['Plans Transferred In']
                plans_transferred_out = row['Plans Transferred Out']
                plans_transferred_in_dollars = row['Plans Transferred In (Dollars)']
                plans_transferred_out_dollars = row['Plans Transferred Out (Dollars)']
                suspended = row['Suspended']
                suspended_dollars = row['Suspended (Dollars)']
                resumed = row['Resumed']
                resumed_dollars = row['Resumed (Dollars)']
                plans_terminated = row['Plans Terminated']
                plans_terminated_dollars = row['Plans Terminated (Dollars)']
                plans_discontinued = row['Plans Discontinued']
                cc_declining_expired = row['CC Declining + CC Expired']
                expired_members = row['Expired Members']
                members_per_day = row['Members per day']
                revenue = row['Revenue']
                pass_wash_percentage = row['Pass Wash Percentage']
                churn_rate = row['Churn Rate']

                if ChurnByDay.get(site) != None:
                    ChurnByDay[site].append(churn_rate)
                else:
                    ChurnByDay[site] = []
                    ChurnByDay[site].append(churn_rate)

        CorporateChurnRates = []
        CorporateChurn = 0
        for locationName, ratesList in ChurnByDay.items():
            if trendsFilePath != None:
                ChurnTotal[locationName] = round(sum(ChurnByDay[locationName])/len(ChurnByDay[locationName]), 4)
                CorporateChurnRates.append(round(sum(ChurnByDay[locationName])/len(ChurnByDay[locationName]), 4))
        
        if CorporateChurnRates != []:
            ChurnTotal[Corporation_Totals_Name] = round(sum(CorporateChurnRates)/len(CorporateChurnRates), 4)


    '''


    #########################################################################################################################################################################
    ############################################################################## XCL LOADING ##############################################################################
    #########################################################################################################################################################################
    #--> Utility Functions to copy a range of cells from one location to another, including their styles and formulas
    def adjust_cell_reference(match, row_shift, col_shift):
        cell_ref = match.group(0)
        col_ref, row_ref = re.match(r"([A-Z]+)([0-9]+)", cell_ref).groups()
        new_col = get_column_letter(column_index_from_string(col_ref) + col_shift)
        new_row = str(int(row_ref) + row_shift)
        return new_col + new_row

    def adjust_formula(formula, row_shift, col_shift):
        pattern = r"(\$?[A-Z]+\$?[0-9]+)"
        return re.sub(pattern, lambda match: adjust_cell_reference(match, row_shift, col_shift), formula)

    def copy_cells(ws, start_row, end_row, start_col, end_col, target_start_row, target_start_col, target_ws=None):
        row_shift = target_start_row - start_row
        col_shift = target_start_col - start_col
        
        # If no target worksheet is given, default to the current worksheet
        if target_ws is None:
            target_ws = ws

        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)):
            for j, cell in enumerate(row):
                target_cell = target_ws.cell(row=target_start_row + i, column=target_start_col + j)
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Adjust the formula
                    adjusted_formula = adjust_formula(cell.value[1:], row_shift, col_shift)
                    target_cell.value = f"={adjusted_formula}"
                else:
                    target_cell.value = cell.value

                # Copy cell style if it has one
                if cell.has_style:
                    target_cell._style = copy(cell._style)
                
                # Copy font, border, fill, number format, etc.
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)


    #--> Populate Worksheet Function to fill in each location block
    def populate_salesmix_worksheet(worksheet, start_row, start_col, location):
        global monthsInReport

        #--> Assign location name as the new block name
        if location != 'Query Server':
            worksheet.cell(row=start_row, column=start_col, value=location)
        else:
            worksheet.cell(row=start_row, column=start_col, value="E-Commerce Website")
        worksheet.cell(row=start_row, column=start_col+5, value=monthsInReport)
        print(AccurateMemberCountAdjuster)
        worksheet.cell(row=start_row, column=start_col+6, value=AccurateMemberCountAdjuster[location])
        worksheet.cell(row=start_row, column=start_col+7, value=HIST_AccurateMemberCountAdjuster[location])
        
        #--> BLOCK: Retail Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            worksheet.cell(row=start_row+3+i, column=start_col+5, value=pkgProps['Price'])
        
        #--> BLOCK: Monthly Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        #### PRESENT PERIOD ENTRIES
        for i, (washPkg, pkgProps) in enumerate(MONTHLY_STATS[location].items()):
            worksheet.cell(row=start_row+13+i, column=start_col+1, value=pkgProps['Quantity'])
            worksheet.cell(row=start_row+13+i, column=start_col+7, value=pkgProps['Estimated Member Count'])
            worksheet.cell(row=start_row+13+i, column=start_col+10, value=pkgProps['Gross New Members'])
        if ChurnTotal.get(location):
            worksheet.cell(row=start_row+18, column=start_col+6, value=ChurnTotal[location])
        else:
            worksheet.cell(row=start_row+18, column=start_col+6, value='Not Available')
        worksheet.cell(row=start_row+18, column=start_col+1, value=MONTHLY_SALES[location]['NET Sales'])
        #### HISTORICAL PERIOD ENTRIES
        for i, (washPkg, pkgProps) in enumerate(HIST_MONTHLY_STATS[location].items()):
            worksheet.cell(row=start_row+13+i, column=start_col+2, value=pkgProps['Quantity'])
            worksheet.cell(row=start_row+13+i, column=start_col+8, value=pkgProps['Estimated Member Count'])
            #worksheet.cell(row=start_row+12+i, column=start_col+6, value=pkgProps['Gross New Members'])
        '''if ChurnTotal.get(location):
            worksheet.cell(row=start_row+17, column=start_col+4, value=ChurnTotal[location])
        else:
            worksheet.cell(row=start_row+17, column=start_col+4, value='Not Available')'''
        worksheet.cell(row=start_row+18, column=start_col+2, value=HIST_MONTHLY_SALES[location]['NET Sales'])

        #--> BLOCK: Combined Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        #### PRESENT DATA ENTRIES
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            if location != 'Query Server':
                worksheet.cell(row=start_row+23+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+23+i, column=start_col+5, value=pkgProps['Price'])
            else:
                worksheet.cell(row=start_row+5+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+5+i, column=start_col+5, value=pkgProps['Price'])
                worksheet.cell(row=start_row+23+i, column=start_col+1, value=pkgProps['Quantity']+MONTHLY_STATS['Query Server'][washPkg]['Quantity'])
                worksheet.cell(row=start_row+23+i, column=start_col+5, value=pkgProps['Price'])
        
        worksheet.cell(row=start_row+28, column=start_col+1, value=COMBINED_SALES[location]['NET Sales'])
        worksheet.cell(row=start_row+28, column=start_col+5, value=COMBINED_SALES[location]['Discounts'])
        #### HISTORICAL DATA ENTRIES
        for i, (washPkg, pkgProps) in enumerate(HIST_COMBINED_STATS[location].items()):
            if location != 'Query Server':
                worksheet.cell(row=start_row+23+i, column=start_col+2, value=pkgProps['Quantity'])
            else:
                worksheet.cell(row=start_row+5+i, column=start_col+2, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+23+i, column=start_col+2, value=pkgProps['Quantity']+HIST_MONTHLY_STATS['Query Server'][washPkg]['Quantity'])
        
        worksheet.cell(row=start_row+28, column=start_col+2, value=HIST_COMBINED_SALES[location]['NET Sales'])
        
        #--> Query Server Modifications:
        if location == "Query Server":
            worksheet.cell(row=start_row+12, column=start_col+1, value="ARM Plans Sold (Pres.)")
            worksheet.cell(row=start_row+12, column=start_col+2, value="ARM Plans Sold (Prev.)")
            worksheet.cell(row=start_row+17, column=start_col, value="Total Plans Sold")
            worksheet.cell(row=start_row+22, column=start_col+1, value="Items Sold (Pres.)")
            worksheet.cell(row=start_row+22, column=start_col+2, value="Items Sold (Prev.)")
            worksheet.cell(row=start_row+27, column=start_col, value="Total Items Sold for Period")
        
        
        createMiniBarGraph(worksheet, start_col, start_row)
        createMiniBarGraph(worksheet, start_col, start_row+20)

        ## GRAPH CREATION!
        #--> Set up the sheet
        if location not in [Corporation_Totals_Name, 'Query Server']:
            ws = wb['Sales Mix by Location (Visual)']
            visualWorksheet = wb.copy_worksheet(ws)
            visualWorksheet.title = f'Sales Mix Charts ({location.strip("wash*u - ")})'

            copy_cells(wb['Sales Mix by Location'], start_row, start_row+rowsToCopy, start_col, start_col+columnsToCopy, 1, 1, visualWorksheet)

            # Generate Graphs ==> Data Reference, Category Reference
            # Retail Sales Mix (Qty)
            createPieChart(visualWorksheet, "M2", Reference(visualWorksheet, min_col=2, min_row=4, max_col=2, max_row=7), Reference(visualWorksheet, min_col=1, min_row=4, max_col=1, max_row=7), "Retail Sales Mix (Qty)")
            # Monthly Membership Mix (Qty)
            createPieChart(visualWorksheet, "O2", Reference(visualWorksheet, min_col=5, min_row=14, max_col=5, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Monthly Membership Mix (Qty)")
            # Combined Wash Mix (Qty)
            createPieChart(visualWorksheet, "Q2", Reference(visualWorksheet, min_col=2, min_row=24, max_col=2, max_row=27), Reference(visualWorksheet, min_col=1, min_row=24, max_col=1, max_row=27), "Combined Wash Mix (Qty)")
            
            # Net Revenue Breakdown (%)
            createPieChart(visualWorksheet, "M19", Reference(visualWorksheet, min_col=3, min_row=41, max_col=3, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown (%)", 13.95, 14.25)
            # Net Revenue Breakdown ($)
            createPieChart(visualWorksheet, "P19", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown ($)", 13.95, 14.25)
            # Retail Revenue Breakdown (%)
            createPieChart(visualWorksheet, "M43", Reference(visualWorksheet, min_col=4, min_row=41, max_col=4, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown (%)", 13.95, 14.25)
            # Retail Revenue Breakdown ($)
            createPieChart(visualWorksheet, "P43", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown ($)", 13.95, 14.25)
            
            # Monthly Membership Utilization
            createBarGraph(visualWorksheet, "T2", Reference(visualWorksheet, min_col=7, min_row=14, max_col=7, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Avg. Monthly Pass Utilization", y_axisTitle="Average # of Washes over Period")
            # Monthly Membership Count
            createBarGraph(visualWorksheet, "T19", Reference(visualWorksheet, min_col=8, min_row=14, max_col=8, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Total Number of Monthly Members", y_axisTitle="Number of Members")
            # Monthly NEW members
            createBarGraph(visualWorksheet, "T36", Reference(visualWorksheet, min_col=11, min_row=14, max_col=11, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Average Num. of NEW Monthly Members", y_axisTitle="Number of Members")

            # Traffic Pie in pie Chart (Retail + Monthly, then monthly broken into its own pie chart)
            createPieInPieChart(visualWorksheet, "T36", Reference(visualWorksheet, min_col=2, min_row=52, max_col=2, max_row=59), Reference(visualWorksheet, min_col=1, min_row=52, max_col=1, max_row=59), "Retail & Monthly Wash Mix", y_axisTitle="Number of Members")

            createMiniBarGraph(visualWorksheet, 1, 3-2)
            createMiniBarGraph(visualWorksheet, 1, 23-2)
        elif location == Corporation_Totals_Name:
            # Due to the nature of the corporate tab, data is built in this function and then churned into graphs, rather than allowing the spreadsheet to handle it.
            # This is because the number of locations is dynamic and changing therefore we cannot hard-code the corporate graph data functions like we do with the others.

            # Copy the template to the workbook
            ws = wb['Sales Mix by Location (Visual)']
            visualWorksheet = wb.copy_worksheet(ws)
            visualWorksheet.title = 'Sales Mix Charts (Corporation)'

            # Position the Corporation Sheet towards the front
            move_worksheet_to_position(wb, visualWorksheet, 2)

            copy_cells(wb['Sales Mix by Location'], start_row, start_row+rowsToCopy, start_col, start_col+columnsToCopy, 1, 1, visualWorksheet)

            # Generate Graphs ==> Data Reference, Category Reference
            # Retail Sales Mix (Qty)
            createPieChart(visualWorksheet, "M2", Reference(visualWorksheet, min_col=2, min_row=4, max_col=2, max_row=7), Reference(visualWorksheet, min_col=1, min_row=4, max_col=1, max_row=7), "Retail Sales Mix (Qty)")
            # Monthly Membership Mix (Qty)
            createPieChart(visualWorksheet, "O2", Reference(visualWorksheet, min_col=5, min_row=14, max_col=5, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Monthly Membership Mix (Qty)")
            # Combined Wash Mix (Qty)
            createPieChart(visualWorksheet, "Q2", Reference(visualWorksheet, min_col=2, min_row=24, max_col=2, max_row=27), Reference(visualWorksheet, min_col=1, min_row=24, max_col=1, max_row=27), "Combined Wash Mix (Qty)")
            
            # Net Revenue Breakdown (%)
            createPieChart(visualWorksheet, "M19", Reference(visualWorksheet, min_col=3, min_row=41, max_col=3, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown (%)", 13.95, 14.25)
            # Net Revenue Breakdown ($)
            createPieChart(visualWorksheet, "P19", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown ($)", 13.95, 14.25)
            # Retail Revenue Breakdown (%)
            createPieChart(visualWorksheet, "M43", Reference(visualWorksheet, min_col=4, min_row=41, max_col=4, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown (%)", 13.95, 14.25)
            # Retail Revenue Breakdown ($)
            createPieChart(visualWorksheet, "P43", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown ($)", 13.95, 14.25)
            
            # Monthly Membership Utilization
            createBarGraph(visualWorksheet, "T2", Reference(visualWorksheet, min_col=7, min_row=14, max_col=7, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Avg. Monthly Pass Utilization", y_axisTitle="Average # of Washes over Period")
            # Monthly Membership Count
            createBarGraph(visualWorksheet, "T19", Reference(visualWorksheet, min_col=8, min_row=14, max_col=8, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Total Number of Monthly Members", y_axisTitle="Number of Members")
            # Monthly NEW members
            createBarGraph(visualWorksheet, "T36", Reference(visualWorksheet, min_col=11, min_row=14, max_col=11, max_row=17), Reference(visualWorksheet, min_col=1, min_row=14, max_col=1, max_row=17), "Average Num. of NEW Monthly Members", y_axisTitle="Number of Members")

            # Traffic Pie in pie Chart (Retail + Monthly, then monthly broken into its own pie chart)
            createPieInPieChart(visualWorksheet, "T36", Reference(visualWorksheet, min_col=2, min_row=52, max_col=2, max_row=59), Reference(visualWorksheet, min_col=1, min_row=52, max_col=1, max_row=59), "Retail & Monthly Wash Mix", y_axisTitle="Number of Members")

            createMiniBarGraph(visualWorksheet, 1, 3-2)
            createMiniBarGraph(visualWorksheet, 1, 23-2)

            # Generate Competitive Graphs (These will compare sites against other sites on basic revenue, traffic, discounts, etc.)
            # Create Data Tables (will be used to input the data into the visuals sheet for the corporation, tracking the cell inputs with python we can dynamically generate & gather graph data.)
            retail_quantity_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in retail_quantity_mix_data: 
                            retail_quantity_mix_data[location] = {}
                        retail_quantity_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Quantity'] - MONTHLY_STATS[location][washPkg]['Quantity']})
            

            retail_revenue_mix_data = {}
            total_retail_revenue_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_retail_revenue_mix_data[location] = COMBINED_SALES[location]['NET Sales'] - MONTHLY_SALES[location]['NET Sales']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in retail_revenue_mix_data: retail_revenue_mix_data[location] = {}
                        retail_revenue_mix_data[location].update({washPkg: (COMBINED_STATS[location][washPkg]['Quantity'] - MONTHLY_STATS[location][washPkg]['Quantity'])*COMBINED_STATS[location][washPkg]['Price']})
                    retail_revenue_mix_data[location].update({'Discounts': COMBINED_SALES[location]['Discounts']})

            monthly_quantity_mix_data = {}

            for location in MONTHLY_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in monthly_quantity_mix_data: 
                            monthly_quantity_mix_data[location] = {}
                        monthly_quantity_mix_data[location].update({washPkg: MONTHLY_STATS[location][washPkg]['Quantity']})
            

            monthly_revenue_mix_data = {}
            total_monthly_revenue_mix_data = {}

            for location in MONTHLY_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_monthly_revenue_mix_data[location] = MONTHLY_SALES[location]['NET Sales']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in monthly_revenue_mix_data: monthly_revenue_mix_data[location] = {}
                        monthly_revenue_mix_data[location].update({washPkg: MONTHLY_STATS[location][washPkg]['Amount']})
            

            combined_quantity_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in combined_quantity_mix_data: 
                            combined_quantity_mix_data[location] = {}
                        combined_quantity_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Quantity']})
            

            corporate_churn_rates_data = {}

            for location in ChurnTotal:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    if location not in corporate_churn_rates_data: 
                        corporate_churn_rates_data[location] = 0
                    corporate_churn_rates_data.update({location: ChurnTotal[location]})
            

            combined_revenue_mix_data = {}
            total_combined_revenue_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_combined_revenue_mix_data[location] = COMBINED_SALES[location]['NET Sales'] - COMBINED_SALES[location]['Discounts']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in combined_revenue_mix_data: combined_revenue_mix_data[location] = {}
                        combined_revenue_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Amount']})
            
            start_row = 75
            start_col = 1
            worksheet = wb[f'Sales Mix Charts (Corporation)']

            accounting_style_2 = NamedStyle(
                name='accounting_style_2', 
                number_format='$#,##0.00',  # or you could use 'Accounting'
                alignment=Alignment(horizontal='right')
            )
            percent_style_2 = NamedStyle(name='percent_style_2', number_format='0.00%')

            '''# Loop through each location and its corresponding wash packages
            for location, washPackages in retail_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row, column=start_col+1, value=location.strip("wash*u - "))
                
                # Increment the row to start adding wash packages for this location
                current_row = start_row + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    if start_col == 1:
                        worksheet.cell(row=current_row, column=start_col, value=washPackage)
                    worksheet.cell(row=current_row, column=start_col + 1, value=value)
                    current_row += 1  # Move down a row for the next wash package
                
                # Move to the next column to start adding the next location
                start_col += 1  # Shift 2 columns to the right for the next location'''
            # Loop through each location and its corresponding wash packages
            for location, washPackages in retail_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=75, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location

            createCorporateBarGraph(worksheet, "AB2", Reference(worksheet, min_col=2, min_row=75, max_col=5, max_row=start_row), Reference(worksheet, min_col=1, min_row=76, max_col=1, max_row=start_row), "Retail Wash Sales by Site (QTY)", y_axisTitle="# Washes Sold")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in retail_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_retail_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_retail_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown (%)", showPercents=True)


            start_row += 3  # Shift 2 rows down for the next location
            washLabelRow = start_row
# Loop through each location and its corresponding wash packages
            for location, washPackages in monthly_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            createCorporateBarGraph(worksheet, "AB37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Wash Rdmds by Site (QTY)", y_axisTitle="# Washes Rdmd")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in monthly_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Revenue Breakdown by Site ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_monthly_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_monthly_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Revenue Breakdown (%)", showPercents=True)




# Loop through each location and its corresponding wash packages
            start_row += 3  # Shift 2 rows down for the next location
            washLabelRow = start_row
            for location, washPackages in combined_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location

            createCorporateBarGraph(worksheet, "AB74", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Combined Wash Sales by Site (QTY)", y_axisTitle="# Washes Sold")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in combined_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_combined_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_combined_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN74", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Combined Revenue Breakdown ($)", showPercents=True)


            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in corporate_churn_rates_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=corporate_churn_rates_data[location]).style = percent_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createBarGraph(visualWorksheet, "P71", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=2, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Churn Rate by Location (%)")




    #--> Modify the template and populate all blocks with site data
    # copy_cells(ws, start_row, end_row, start_col, end_col, target_start_row, target_start_col, target_ws=None):
    wb = openpyxl.load_workbook(templateFilePath)
    ws = wb['Sales Mix by Location']

    start_col = 1
    #--> Assign a report date to the report.
    ws.cell(row=1, column=1, value="Report Periods:")
    ws.cell(row=1, column=2, value=complete_date_str + " (Pres.)  &  " + complete_date_str_hist + " (Prev.)")
    locations = list(MONTHLY_STATS.keys())
    locations.remove("Corporation Totals")
    if "Query Server" in locations:
        locations.remove("Query Server")
    for location in locations:
        print(f"Creating {location} Stats Block...")
        copy_cells(ws, 3, rowsToCopy, 1, columnsToCopy, 3, start_col) #copy_cells(ws, 1, 28, 1, 6, 1, start_col)
        populate_salesmix_worksheet(ws, 3, start_col, location)
        start_col += columnsToCopy+1
    
    copy_cells(ws, 3, rowsToCopy, 1, columnsToCopy, rowsToCopy+2, 1)
    populate_salesmix_worksheet(ws, rowsToCopy+2, 1, Corporation_Totals_Name)

    if not "Query Server" in blacklisted_sites:
        print("Creating Query Server Stats Block...")
        copy_cells(ws, 3, rowsToCopy, 1, columnsToCopy, rowsToCopy+2, columnsToCopy+2)
        populate_salesmix_worksheet(ws, rowsToCopy+2, columnsToCopy+2, "Query Server")

    
        
    #--> Discounts Worksheet Stuff
    # Define styles
    int_style = NamedStyle(name='integer_style', number_format='#,##0')
    accounting_style = NamedStyle(
        name='accounting_style', 
        number_format='#,##0.00',  # or you could use 'Accounting'
        alignment=Alignment(horizontal='right')
    )
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    # Modify the populate_discountmix_worksheet function to include formatting
    def populate_discountmix_worksheet(worksheet, start_row, start_col, location, discount_data):
        worksheet.cell(row=start_row, column=start_col, value=location)
        
        # Adding headers
        headers = ["Discount Name", "Quantity Used", "Price*", "Total Discount Dollars", "Discount Distribution (Qty)", "Discount Distribution ($)"]
        for j, header in enumerate(headers):
            worksheet.cell(row=start_row + 2, column=start_col + j, value=header)
            
        total_quantity = 0
        total_amount = 0

        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            total_quantity += discount_info['Quantity'] if discount_info['Quantity'] is not None else 0
            total_amount += discount_info['Amount'] if discount_info['Amount'] is not None else 0
        
        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            if discount_name is not None:

                # Add discount name
                worksheet.cell(row=start_row+3+i, column=start_col, value=discount_name)
                
                # Add and format Quantity
                cell_quantity = worksheet.cell(row=start_row+3+i, column=start_col+1, value=discount_info['Quantity'])
                cell_quantity.style = int_style
                
                # Add and format Price
                cell_price = worksheet.cell(row=start_row+3+i, column=start_col+2, value=discount_info['Price'])
                cell_price.style = accounting_style
                
                # Add and format Amount
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+3, value=discount_info['Amount'])
                cell_amount.style = accounting_style

                # Add and format Distribution by Quantity
                distribution_value = discount_info['Quantity'] / total_quantity if total_quantity != 0 else 0
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+4, value=distribution_value)
                cell_amount.style = percent_style  # Apply the percentage style

                # Add and format Distribution by Amount
                distribution_value = discount_info['Amount'] / total_amount if total_amount != 0 else 0
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+5, value=distribution_value)
                cell_amount.style = percent_style  # Apply the percentage style
                
                
        '''
        # Add and format "Total"
        total_label_cell = worksheet.cell(row=start_row+4+i, column=start_col, value="Total")
        total_label_cell.font = Font(bold=True)

        # Add and format Total Quantity
        cell_total_quantity = worksheet.cell(row=start_row+4+i, column=start_col+1, value=total_quantity)
        cell_total_quantity.style = int_style  # If you have defined int_style elsewhere
        cell_total_quantity.font = Font(bold=True)

        # Add and format Total Amount
        cell_total_amount = worksheet.cell(row=start_row+4+i, column=start_col+3, value="${:.2f}".format(total_amount))
        cell_total_amount.style = accounting_style  # If you have defined accounting_style elsewhere
        cell_total_amount.font = Font(bold=True)
        '''
        
        # Add a Price disclaimer explaining that some prices are averages
        if location == Corporation_Totals_Name:
            disclaim1 = worksheet.cell(row=start_row+4+i+2, column=start_col, value="NOTES")
            disclaim2 = worksheet.cell(row=start_row+4+i+3, column=start_col, value="Price*: Some 'Open Price' items such as 'Misc. Wash Discount' show the average amount that was paid out over the period under that discount item.")
            disclaim3 = worksheet.cell(row=start_row+4+i+4, column=start_col, value="Discounts that are not applied (are missing on the GSR for the location while parsing; thus quantity is 0) will not be shown on the report for that location.")
            disclaim4 = worksheet.cell(row=start_row+4+i+5, column=start_col, value="Due to API limitations with Excel, a proper Total Row for the table cannot be 'enabled' by default. Therefore, it is recommended that you (the end user) enable the total row by clicking anywhere in the table, click table design in the ribbon bar and then check 'Total Row' in the Table Style Options section.")
            disclaim1.font = Font(bold=True)
            disclaim2.font = Font(bold=True)
            disclaim3.font = Font(bold=True)
            disclaim4.font = Font(bold=True)

        tableName = re.sub(r'\W+', '', location.replace(" ", ""))
        # Adding table and table banding
        tab = Table(displayName=tableName, ref=f"{worksheet.cell(row=start_row + 2, column=start_col).coordinate}:{worksheet.cell(row=start_row + 4 + i, column=start_col + 5).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium23", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    # Read the input CSV to a DataFrame
    # TODO: Historical GSR Discounts need to be read and added to a dictionary
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    # Aggregate the data
    discount_data = {}
    miscDisctAverages = {}
    discountTotals = {}
    discount_data[Corporation_Totals_Name] = {}
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        discount_name = row['Item Name']
        quantity = row['Quantity']
        price = row['Price']
        amount = row['Amount']

        if (category in Discount_Categories and 
            site not in blacklisted_sites and 
            pd.notna(discount_name) and 
            pd.notna(quantity) and 
            pd.notna(price) and 
            pd.notna(amount)):
            
            if site not in discount_data:
                discount_data[site] = {}
                
            if discount_name not in discount_data[site]:
                discount_data[site][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            if discount_name not in discount_data[Corporation_Totals_Name]:
                discount_data[Corporation_Totals_Name][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            discount_data[site][discount_name]['Quantity'] += quantity
            discount_data[Corporation_Totals_Name][discount_name]['Quantity'] += quantity
            if discount_name not in miscDisctAverages:
                miscDisctAverages[discount_name] = []
            miscDisctAverages[discount_name].append(price if pd.notna(price) else 0)
            discount_data[site][discount_name]['Amount'] += amount
            discount_data[Corporation_Totals_Name][discount_name]['Amount'] += amount
    
    # Average the prices
    # Used for Misc. discounts to see the average amount of money paid out to customers
    locations = list(discount_data.keys())

    for location in locations:
        for discount in discount_data[location]:
            try:
                discount_data[location][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
                discount_data[Corporation_Totals_Name][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
            except:
                discount_data[location][discount]['Price'] = 0

    # Load the template workbook
    ws = wb['Discount Mix by Location']
    # Initial starting column for the first location
    start_col = 1

    # Populate the worksheet
    for location, data in discount_data.items():
        copy_cells(ws, 1, 3, 1, 6, 1, start_col)
        populate_discountmix_worksheet(ws, 1, start_col, location, data)
        start_col += 7
    
    #########################################################################################################################################################################
    ############################################################################ GRAPHS CREATION ############################################################################
    #########################################################################################################################################################################
        
    wb._sheets.remove(wb['Sales Mix by Location (Visual)'])
    '''
    MONTHLY_STATS = {}
    COMBINED_STATS = {}
    COMBINED_SALES = {}
    MONTHLY_SALES = {}

    HIST_MONTHLY_STATS = {}
    HIST_COMBINED_STATS = {}
    HIST_COMBINED_SALES = {}
    HIST_MONTHLY_SALES = {}
    '''

    wbName = f"Sales Mix - {short_date_str}"
    return (
        wb, 
        wbName, 
        {
            "Monthly Stats": MONTHLY_STATS,
            "Combined Stats": COMBINED_STATS,
            "Combined Sales": COMBINED_SALES,
            "Monthly Sales": MONTHLY_SALES,
            "Historical Monthly Stats": HIST_MONTHLY_STATS,
            "Historical Combined Stats": HIST_COMBINED_STATS,
            "Historical Combined Sales": HIST_COMBINED_SALES,
            "Historical Monthly Sales": HIST_MONTHLY_SALES,
            "Churn Data Dictionary": ChurnDataDictionary,
            "Historical Churn Data Dictionary": HIST_ChurnDataDictionary,
            "Churn Total": ChurnTotal,
            "Historical Churn Total": HIST_ChurnTotal
        },
        complete_date_str + " (Pres.)  &  " + complete_date_str_hist + " (Prev.)"
    )


    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################


def createSalesMixSheet(gsrFilePath, templateFilePath, fileNameForParser, trendsFilePath=None, excludedLocations=None):
    #--> Create Dictionaries for looking up item names.
    # TODO: This needs to be attached to a database and allow insertions through the manager portal or a form.
    global monthsInReport
    monthsInReport = 1
    MONTHLY_STATS = {}
    COMBINED_STATS = {}
    COMBINED_SALES = {}
    MONTHLY_SALES = {}


    #--> Utility Functions
    def find_parent(json_obj, target_str, start_parent=None, current_parent=None):
        """
        Searches for a string in a nested JSON object (Python dict) starting from a specified parent,
        and returns the parent key under which the string resides.
        
        Parameters:
        - json_obj (dict): The JSON object to search through.
        - target_str (str): The string to look for.
        - start_parent (str): The parent key from where to start the search.
        - current_parent: The current parent key, used for recursion.
        
        Returns:
        - The parent key under which the string resides, or None if the string is not found.
        """
        if start_parent and current_parent is None:
            json_obj = json_obj.get(start_parent, {})
            
        for key, value in json_obj.items():
            if key == target_str:
                return current_parent
            if isinstance(value, dict):
                result = find_parent(value, target_str, None, key)
                if result:
                    return result
            elif isinstance(value, (list, set)):
                if target_str in value:
                    return key
        return None

    def move_worksheet_to_position(workbook, worksheet, position):
        """
        Move a worksheet to a specific position within a workbook.

        Parameters:
        - workbook: The Workbook object where the worksheet resides.
        - worksheet: The Worksheet object to be moved.
        - position: The position (0-based index) where the worksheet should be moved to.

        Returns:
        - True if successful, False otherwise.
        """
        if worksheet not in workbook._sheets:
            print("The worksheet is not in the given workbook.")
            return False

        if position < 0 or position >= len(workbook._sheets):
            print("Invalid position.")
            return False

        # Remove the original worksheet from the list
        workbook._sheets.remove(worksheet)

        # Insert it back at the specified position
        workbook._sheets.insert(position, worksheet)

        return True
    
    #--> Utility Functions (Graphs)
    def createMiniBarGraph(worksheet, START_COL, START_ROW):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Set the data range for the chart
        # Assuming data for the chart is in cells A4:B7
        values = Reference(ws, min_col=START_COL+1, min_row=START_ROW+3, max_col=START_COL+1, max_row=START_ROW+6)
        labels = Reference(ws, min_col=START_COL, min_row=START_ROW+3, max_row=START_ROW+6)
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(labels)

        # Set the dimensions of the chart
        chart.height = 4.05  # height
        chart.width = 11.35  # width

        # Remove the legend
        chart.legend = None

        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

        # Change bar colors (optional)
        colors = ['4682B4']  # Blue
        for i, s in enumerate(chart.series):
            s.graphicalProperties.solidFill = colors[i % len(colors)]
        
        CHART_PLACEMENT_COL = START_COL + 4
        CHART_PLACEMENT_ROW = START_ROW + 2

        # Add the chart to the worksheet
        ws.add_chart(chart, f"{get_column_letter(CHART_PLACEMENT_COL)}{CHART_PLACEMENT_ROW}")
    
    def createPieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None):
        ws = worksheet

        # Create a bar chart
        chart = PieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 9.5 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showLeaderLines = True


        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    def createCorporatePieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, showPercents=False):
        ws = worksheet

        # Create a bar chart
        chart = PieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 20
        chart.width = chartWidth if chartWidth != None else 20 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showLeaderLines = True
        if showPercents == True:
            chart.dLbls.showPercent = True


        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle

    def createBarGraph(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 19 


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showLeaderLines = True
        chart.legend = None
        chart.y_axis.title = y_axisTitle if y_axisTitle != None else ''
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle

    def createCorporateBarGraph(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None, useDataTitle=True):
        ws = worksheet

        # Create a bar chart
        chart = BarChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=useDataTitle)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 20
        chart.width = chartWidth if chartWidth != None else 20


        # Add data labels
        chart.dLbls = DataLabelList()
        #chart.dLbls.showVal = True
        #chart.dLbls.showLeaderLines = True
        #chart.legend = None
        chart.y_axis.title = y_axisTitle if y_axisTitle != None else ''

        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    def createPieInPieChart(worksheet, placementCell, dataReference, categoryReference, chartTitle, chartHeight=None, chartWidth=None, valueStyle=None, y_axisTitle=None):
        ws = worksheet

        # Create a bar chart
        chart = ProjectedPieChart()

        # Add the data range for the chart
        chart.add_data(dataReference, titles_from_data=False)
        chart.set_categories(categoryReference)

        # Set the dimensions of the chart
        chart.height = chartHeight if chartHeight != None else 9.3
        chart.width = chartWidth if chartWidth != None else 19

        chart.splitPos=4


        # Add data labels
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showLeaderLines = True
        

        # Add the chart to the worksheet
        ws.add_chart(chart, placementCell)

        chart.title = chartTitle
    
    




    #########################################################################################################################################################################
    ############################################################################## GSR PARSING ##############################################################################
    #########################################################################################################################################################################
    #--> Gather Start/End Dates for the report

    # Read the CSV file to extract the start date and end date
    df = pd.read_csv(gsrFilePath, nrows=2)  # Reading only the first 2 rows

    def extract_dates_from_filepath():
        # Extract the file name from the file path
        filename = fileNameForParser
        global monthsInReport
        # Use regular expression to find dates in the filename
        match = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})", filename)
        
        if match:
            start_date_str, end_date_str = match.groups()
            
            # Convert the extracted date strings to datetime objects
            start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d')

            start_date = start_date_dt.date()
            end_date = end_date_dt.date()

            # Convert the start and end dates from strings to datetime objects
            # Calculate the difference in days between the two dates
            delta = (end_date_dt- start_date_dt).days

            # Calculate the number of months
            monthsInReport = (delta / 30.44) if (delta / 30.44) > 1 else 1  # On average, a month is about 30.44 days, always make sure it is greater than 1 or else there will be in correct stats

            return start_date, end_date
        else:
            return None, None

    # Example usage
    start_date, end_date = extract_dates_from_filepath()
    print(start_date, end_date)
    # Format the date strings
    short_date_str, complete_date_str = str(start_date) + ' - ' + str(end_date), str(start_date) + ' - ' + str(end_date)
        
    #--> Prepare dictionaries for data entry
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        print("Calculating for", site)
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']

        if site not in blacklisted_sites:
            if site not in MONTHLY_STATS:
                MONTHLY_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if site not in COMBINED_STATS:
                COMBINED_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            if Corporation_Totals_Name not in MONTHLY_STATS:
                MONTHLY_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if Corporation_Totals_Name not in COMBINED_STATS:
                COMBINED_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            
            if site not in MONTHLY_SALES:
                MONTHLY_SALES[site] = {"NET Sales": 0}
            if site not in COMBINED_SALES:
                COMBINED_SALES[site] = {"Discounts": 0, "NET Sales": 0}
            if Corporation_Totals_Name not in MONTHLY_SALES:
                MONTHLY_SALES[Corporation_Totals_Name] = {"NET Sales": 0}
            if Corporation_Totals_Name not in COMBINED_SALES:
                COMBINED_SALES[Corporation_Totals_Name] = {"Discounts": 0, "NET Sales": 0}

    #--> Gather TOTAL sales & discounts details
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            COMBINED_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
            COMBINED_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            if category in Discount_Categories:
                COMBINED_SALES[site]['Discounts'] += amount if pd.isna(amount) != True else 0
                COMBINED_SALES[Corporation_Totals_Name]['Discounts'] += amount if pd.isna(amount) != True else 0

            # Handle the Query Server
            if site == 'Query Server':
                if find_parent(Website_PKG_Names, item_name, start_parent="Monthly") != None or category in Monthly_Total_Categories:
                    MONTHLY_SALES['Query Server']['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            # Any other locations
            else: 
                if category in Monthly_Total_Categories:
                    MONTHLY_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0

            

    #--> Gather Sales Mix
    """
    Calculating Combined, Monthly, and Retail Stats:

    Components:
    - ARM Redeemer: A package leader item triggered when a FastPass Plan customer pays.
    - ARM Plan Packages: Package leader items contained within the ARM Redeemer.
    - Wash Package: Actual wash service (e.g., 'Express', 'Protect').
    - ARM Redeemer Discount: Makes the Wash Package cost $0 for FastPass Plan customers.

    Transaction Flow:
    1. FastPass Plan customer arrives, triggering the ARM Redeemer.
    2. The ARM Redeemer contains ARM Plan Packages, each of which includes:
        1. A Wash Package (e.g., 'Express', 'Protect').
        2. An ARM Redeemer Discount, setting the Wash Package cost to $0.

    Stat Calculations:
    - Combined Stats (Retail + Monthly): Both retail and monthly transactions are recorded under 'Basic Washes' because they both trigger a Wash Package.
    - Monthly Stats: Tracked when an ARM Redeemer is activated (exclusive to monthly FastPass Plan customers).
    - Retail Stats: Obtained by subtracting Monthly Stats from Combined Stats.

    Formula:
    Retail Stats = Combined Stats - Monthly Stats

    Template:
    The template is Excel based and therefore can be set up early to calculate the retail data based on the input from this script for combined and monthly.
    """
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    MonthlyAmounts = {}
    CombinedAmounts = {}
    AccurateMemberCountAdjuster = {}
    ChurnDataDictionary = {}
    ChurnTotal  = {}
    price_sum_dict = {}  # This will hold the sum of prices for each item for each site
    price_count_dict = {}  # This will hold the count of entries for each item for each site
    totalCorporateDiscont = 0

    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        count = row['Count']
        price = row['Price']
        quantity = row['Quantity']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            if ChurnDataDictionary.get(site) == None:
                ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            if ChurnDataDictionary.get(Corporation_Totals_Name) == None:
                ChurnDataDictionary[Corporation_Totals_Name] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            
            if AccurateMemberCountAdjuster.get(site) == None:
                AccurateMemberCountAdjuster[site] = 0
            if item_name in ['WEB Discontinue ARM', 'Discontinue ARM Plan']:
                if AccurateMemberCountAdjuster.get(site) != None:
                    AccurateMemberCountAdjuster[site] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    totalCorporateDiscont += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    ChurnDataDictionary[site]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                    ChurnDataDictionary[Corporation_Totals_Name]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Recharge_Names, item_name) # Member Count = ARM Recharges
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(Website_PKG_Names, item_name, "Monthly") # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld | Gross New Members = ARM Sold
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[site][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                 
            washPkg = find_parent(ARM_Termination_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld - ARM Terminations
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[site][washPkg]['Estimated Member Count'])
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'])
                ChurnDataDictionary[site]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
            
            if item_name == "Switch ARM Plan":
                ChurnDataDictionary[site]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.
                ChurnDataDictionary[Corporation_Totals_Name]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Qty is already negative in the GSR, need to add it.

            # Handle the Query Server
            if site == 'Query Server':
                # We're only going to show the amount of passes sold rather than the redemption for the
                # Query Server because you cannot redeem passes online.
                washPkg = find_parent(Website_PKG_Names, item_name, "Monthly")
                if washPkg != None:
                    MONTHLY_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members
    
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0)
                
                washPkg = find_parent(Website_PKG_Names, item_name, "Retail")
                if washPkg != None:
                    COMBINED_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        CombinedAmounts[washPkg] = []
                    CombinedAmounts[washPkg].append(price if pd.isna(price) != True else 0)

            # Any other locations
            else: 
                washPkg = find_parent(ARM_PKG_Names, item_name)
                # Calculate MONTHLY Stats using the redemption items
                if category in ["ARM Plans Redeemed", "Club Plans Redeemed"] and washPkg is not None:
                    MONTHLY_STATS[site][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[site][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[site][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS[site][washPkg]['Price'] += abs(price) if pd.isna(price) != True else 0

                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0) 
                elif category in ['Basic Washes'] and item_name is not None and item_name in COMBINED_STATS[site]:
                    COMBINED_STATS[site][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[site][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[site][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS[site][item_name]['Price'] += amount/quantity if pd.isna(quantity) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in CombinedAmounts:
                        CombinedAmounts[item_name] = []
                    CombinedAmounts[item_name].append(price if pd.isna(price) != True else 0)
                    '''if not site in price_sum_dict:
                        price_sum_dict[site] = {}
                    if not item_name in price_sum_dict[site]:
                        price_sum_dict[site][item_name] = []
                    price_sum_dict[site][item_name].append(price if pd.isna(price) != True else 0)'''

    for washPkg in ARM_Sold_Names:
        COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount']/COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount']/MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        #COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(CombinedAmounts[washPkg])/len(CombinedAmounts[washPkg])
        #MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(MonthlyAmounts[washPkg])/len(MonthlyAmounts[washPkg])
    AccurateMemberCountAdjuster[Corporation_Totals_Name] = totalCorporateDiscont

    for location, stats in ChurnDataDictionary.items():
        # ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
        ChurnTotal[location] = (ChurnDataDictionary[location]["Terminated"]+ChurnDataDictionary[location]["Discontinued"]) / (ChurnDataDictionary[location]["Recharged"]+ChurnDataDictionary[location]["Sold"]+ChurnDataDictionary[location]["Discontinued"]-ChurnDataDictionary[location]["Switched"])
    
    #########################################################################################################################################################################
    ############################################################################## T&C PARSING ##############################################################################
    #########################################################################################################################################################################
    '''
        This can be used to extract the average churn over the period for a trends and comparisons report. The stats from this were off compared to SUDS so it was decided that we would use the 
        SUDS Formula to determine the Churn Rate using the items from the GSR alone. This eliminates the need for the Trends and Comparisons report.
        Formula we will use: Churn Rate = (Terminated + Discontinued) / (Recharged + Sold + Discontinued - Switched)


        ChurnByDay = {}
        ChurnTotal = {}
        if trendsFilePath is not None:
            df = pd.read_csv(trendsFilePath)
            for _, row in df.iterrows():
                date = row['Date']
                plan = row['Plan']
                site = row['Site']
                washes_per_day = row['Washes per day']
                upsells_dollars = row['Upsells (Dollars)']
                upsells = row['Upsells']
                plans_sold_dollars = row['Plans Sold (Dollars)']
                plans_sold = row['Plans Sold']
                recharges_dollars = row['Recharges (Dollars)']
                recharges = row['Recharges']
                plans_transferred_in = row['Plans Transferred In']
                plans_transferred_out = row['Plans Transferred Out']
                plans_transferred_in_dollars = row['Plans Transferred In (Dollars)']
                plans_transferred_out_dollars = row['Plans Transferred Out (Dollars)']
                suspended = row['Suspended']
                suspended_dollars = row['Suspended (Dollars)']
                resumed = row['Resumed']
                resumed_dollars = row['Resumed (Dollars)']
                plans_terminated = row['Plans Terminated']
                plans_terminated_dollars = row['Plans Terminated (Dollars)']
                plans_discontinued = row['Plans Discontinued']
                cc_declining_expired = row['CC Declining + CC Expired']
                expired_members = row['Expired Members']
                members_per_day = row['Members per day']
                revenue = row['Revenue']
                pass_wash_percentage = row['Pass Wash Percentage']
                churn_rate = row['Churn Rate']

                if ChurnByDay.get(site) != None:
                    ChurnByDay[site].append(churn_rate)
                else:
                    ChurnByDay[site] = []
                    ChurnByDay[site].append(churn_rate)

        CorporateChurnRates = []
        CorporateChurn = 0
        for locationName, ratesList in ChurnByDay.items():
            if trendsFilePath != None:
                ChurnTotal[locationName] = round(sum(ChurnByDay[locationName])/len(ChurnByDay[locationName]), 4)
                CorporateChurnRates.append(round(sum(ChurnByDay[locationName])/len(ChurnByDay[locationName]), 4))
        
        if CorporateChurnRates != []:
            ChurnTotal[Corporation_Totals_Name] = round(sum(CorporateChurnRates)/len(CorporateChurnRates), 4)


    '''


    #########################################################################################################################################################################
    ############################################################################## XCL LOADING ##############################################################################
    #########################################################################################################################################################################
    #--> Utility Functions to copy a range of cells from one location to another, including their styles and formulas
    def adjust_cell_reference(match, row_shift, col_shift):
        cell_ref = match.group(0)
        col_ref, row_ref = re.match(r"([A-Z]+)([0-9]+)", cell_ref).groups()
        new_col = get_column_letter(column_index_from_string(col_ref) + col_shift)
        new_row = str(int(row_ref) + row_shift)
        return new_col + new_row

    def adjust_formula(formula, row_shift, col_shift):
        pattern = r"(\$?[A-Z]+\$?[0-9]+)"
        return re.sub(pattern, lambda match: adjust_cell_reference(match, row_shift, col_shift), formula)

    def copy_cells(ws, start_row, end_row, start_col, end_col, target_start_row, target_start_col, target_ws=None):
        row_shift = target_start_row - start_row
        col_shift = target_start_col - start_col
        
        # If no target worksheet is given, default to the current worksheet
        if target_ws is None:
            target_ws = ws

        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)):
            for j, cell in enumerate(row):
                target_cell = target_ws.cell(row=target_start_row + i, column=target_start_col + j)
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Adjust the formula
                    adjusted_formula = adjust_formula(cell.value[1:], row_shift, col_shift)
                    target_cell.value = f"={adjusted_formula}"
                else:
                    target_cell.value = cell.value

                # Copy cell style if it has one
                if cell.has_style:
                    target_cell._style = copy(cell._style)
                
                # Copy font, border, fill, number format, etc.
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)


    #--> Populate Worksheet Function to fill in each location block
    def populate_salesmix_worksheet(worksheet, start_row, start_col, location):
        global monthsInReport

        #--> Assign location name as the new block name
        if location != 'Query Server':
            worksheet.cell(row=start_row, column=start_col, value=location)
        else:
            worksheet.cell(row=start_row, column=start_col, value="E-Commerce Website")
        worksheet.cell(row=start_row, column=start_col+5, value=monthsInReport)
        print(AccurateMemberCountAdjuster)
        worksheet.cell(row=start_row, column=start_col+6, value=AccurateMemberCountAdjuster[location])
        
        #--> BLOCK: Retail Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            worksheet.cell(row=start_row+3+i, column=start_col+3, value=pkgProps['Price'])
        
        #--> BLOCK: Monthly Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(MONTHLY_STATS[location].items()):
            worksheet.cell(row=start_row+12+i, column=start_col+1, value=pkgProps['Quantity'])
            worksheet.cell(row=start_row+12+i, column=start_col+5, value=pkgProps['Estimated Member Count'])
            worksheet.cell(row=start_row+12+i, column=start_col+6, value=pkgProps['Gross New Members'])
        if ChurnTotal.get(location):
            worksheet.cell(row=start_row+17, column=start_col+4, value=ChurnTotal[location])
        else:
            worksheet.cell(row=start_row+17, column=start_col+4, value='Not Available')
        worksheet.cell(row=start_row+17, column=start_col+1, value=MONTHLY_SALES[location]['NET Sales'])

        #--> BLOCK: Combined Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            if location != 'Query Server':
                worksheet.cell(row=start_row+21+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+21+i, column=start_col+3, value=pkgProps['Price'])
            else:
                worksheet.cell(row=start_row+3+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+3+i, column=start_col+3, value=pkgProps['Price'])
                worksheet.cell(row=start_row+21+i, column=start_col+1, value=pkgProps['Quantity']+MONTHLY_STATS['Query Server'][washPkg]['Quantity'])
                worksheet.cell(row=start_row+21+i, column=start_col+3, value=pkgProps['Price'])
        
        worksheet.cell(row=start_row+26, column=start_col+1, value=COMBINED_SALES[location]['NET Sales'])
        worksheet.cell(row=start_row+26, column=start_col+3, value=COMBINED_SALES[location]['Discounts'])
        
        #--> Query Server Modifications:
        if location == "Query Server":
            worksheet.cell(row=start_row+11, column=start_col+1, value="New Passes Sold")
            worksheet.cell(row=start_row+16, column=start_col, value="Total ARM Plans Sold")
            worksheet.cell(row=start_row+20, column=start_col+1, value="Items Sold")
            worksheet.cell(row=start_row+25, column=start_col, value="Total Items Sold for Period")
        
        
        createMiniBarGraph(worksheet, start_col, start_row)
        createMiniBarGraph(worksheet, start_col, start_row+18)

        ## GRAPH CREATION!
        #--> Set up the sheet
        if location not in [Corporation_Totals_Name, 'Query Server']:
            ws = wb['Sales Mix by Location (Visual)']
            visualWorksheet = wb.copy_worksheet(ws)
            visualWorksheet.title = f'Sales Mix Charts ({location.strip("wash*u - ")})'

            copy_cells(wb['Sales Mix by Location'], start_row, start_row+28, start_col, start_col+6, 1, 1, visualWorksheet)

            # Generate Graphs
            # Retail Sales Mix (Qty)
            createPieChart(visualWorksheet, "I2", Reference(visualWorksheet, min_col=3, min_row=4, max_col=3, max_row=7), Reference(visualWorksheet, min_col=1, min_row=4, max_col=1, max_row=7), "Retail Sales Mix (Qty)")
            # Monthly Membership Mix (Qty)
            createPieChart(visualWorksheet, "K2", Reference(visualWorksheet, min_col=3, min_row=13, max_col=3, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Monthly Membership Mix (Qty)")
            # Combined Wash Mix (Qty)
            createPieChart(visualWorksheet, "M2", Reference(visualWorksheet, min_col=3, min_row=22, max_col=3, max_row=25), Reference(visualWorksheet, min_col=1, min_row=22, max_col=1, max_row=25), "Combined Wash Mix (Qty)")
            
            # Net Revenue Breakdown (%)
            createPieChart(visualWorksheet, "I19", Reference(visualWorksheet, min_col=3, min_row=41, max_col=3, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown (%)", 13.95, 14.25)
            # Net Revenue Breakdown ($)
            createPieChart(visualWorksheet, "L19", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown ($)", 13.95, 14.25)
            # Retail Revenue Breakdown (%)
            createPieChart(visualWorksheet, "I43", Reference(visualWorksheet, min_col=4, min_row=41, max_col=4, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown (%)", 13.95, 14.25)
            # Retail Revenue Breakdown ($)
            createPieChart(visualWorksheet, "L43", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown ($)", 13.95, 14.25)
            
            # Monthly Membership Utilization
            createBarGraph(visualWorksheet, "P2", Reference(visualWorksheet, min_col=5, min_row=13, max_col=5, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Avg. Monthly Pass Utilization", y_axisTitle="Average # of Washes over Period")
            # Monthly Membership Count
            createBarGraph(visualWorksheet, "P19", Reference(visualWorksheet, min_col=6, min_row=13, max_col=6, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Total Number of Monthly Members", y_axisTitle="Number of Members")
            # Monthly NEW members
            createBarGraph(visualWorksheet, "P36", Reference(visualWorksheet, min_col=7, min_row=13, max_col=7, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Average Num. of NEW Monthly Members", y_axisTitle="Number of Members")

            # Traffic Pie in pie Chart (Retail + Monthly, then monthly broken into its own pie chart)
            createPieInPieChart(visualWorksheet, "P36", Reference(visualWorksheet, min_col=2, min_row=52, max_col=2, max_row=59), Reference(visualWorksheet, min_col=1, min_row=52, max_col=1, max_row=59), "Retail & Monthly Wash Mix", y_axisTitle="Number of Members")

            createMiniBarGraph(visualWorksheet, 1, 3-2)
            createMiniBarGraph(visualWorksheet, 1, 21-2)
        elif location == Corporation_Totals_Name:
            # Due to the nature of the corporate tab, data is built in this function and then churned into graphs, rather than allowing the spreadsheet to handle it.
            # This is because the number of locations is dynamic and changing therefore we cannot hard-code the corporate graph data functions like we do with the others.

            # Copy the template to the workbook
            ws = wb['Sales Mix by Location (Visual)']
            visualWorksheet = wb.copy_worksheet(ws)
            visualWorksheet.title = 'Sales Mix Charts (Corporation)'

            # Position the Corporation Sheet towards the front
            move_worksheet_to_position(wb, visualWorksheet, 2)

            copy_cells(wb['Sales Mix by Location'], start_row, start_row+28, start_col, start_col+6, 1, 1, visualWorksheet)

            # Generate Graphs (Same graphs as designed from the locations)
            # Retail Sales Mix (Qty)
            createPieChart(visualWorksheet, "I2", Reference(visualWorksheet, min_col=3, min_row=4, max_col=3, max_row=7), Reference(visualWorksheet, min_col=1, min_row=4, max_col=1, max_row=7), "Retail Sales Mix (Qty)")
            # Monthly Membership Mix (Qty)
            createPieChart(visualWorksheet, "K2", Reference(visualWorksheet, min_col=3, min_row=13, max_col=3, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Monthly Membership Mix (Qty)")
            # Combined Wash Mix (Qty)
            createPieChart(visualWorksheet, "M2", Reference(visualWorksheet, min_col=3, min_row=22, max_col=3, max_row=25), Reference(visualWorksheet, min_col=1, min_row=22, max_col=1, max_row=25), "Combined Wash Mix (Qty)")
            
            # Net Revenue Breakdown (%)
            createPieChart(visualWorksheet, "I19", Reference(visualWorksheet, min_col=3, min_row=41, max_col=3, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown (%)", 13.95, 14.25)
            # Net Revenue Breakdown ($)
            createPieChart(visualWorksheet, "L19", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=46), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=46), "Net Revenue Breakdown ($)", 13.95, 14.25)
            # Retail Revenue Breakdown (%)
            createPieChart(visualWorksheet, "I43", Reference(visualWorksheet, min_col=4, min_row=41, max_col=4, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown (%)", 13.95, 14.25)
            # Retail Revenue Breakdown ($)
            createPieChart(visualWorksheet, "L43", Reference(visualWorksheet, min_col=2, min_row=41, max_col=2, max_row=45), Reference(visualWorksheet, min_col=1, min_row=41, max_col=1, max_row=45), "Retail Revenue Breakdown ($)", 13.95, 14.25)
            
            # Monthly Membership Utilization
            createBarGraph(visualWorksheet, "P53", Reference(visualWorksheet, min_col=5, min_row=13, max_col=5, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Avg. Monthly Pass Utilization", y_axisTitle="Average # of Washes")
            # Monthly Membership Count
            createBarGraph(visualWorksheet, "P19", Reference(visualWorksheet, min_col=6, min_row=13, max_col=6, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Average Num. of Monthly Members", y_axisTitle="Number of Members")
            # Monthly NEW members
            createBarGraph(visualWorksheet, "P36", Reference(visualWorksheet, min_col=7, min_row=13, max_col=7, max_row=16), Reference(visualWorksheet, min_col=1, min_row=13, max_col=1, max_row=16), "Average Num. of NEW Monthly Members", y_axisTitle="Number of Members")
        
            # Traffic Pie in pie Chart (Retail + Monthly, then monthly broken into its own pie chart)
            createPieInPieChart(visualWorksheet, "P2", Reference(visualWorksheet, min_col=2, min_row=52, max_col=2, max_row=59), Reference(visualWorksheet, min_col=1, min_row=52, max_col=1, max_row=59), "Retail & Monthly Wash Mix")

            createMiniBarGraph(visualWorksheet, 1, 3-2)
            createMiniBarGraph(visualWorksheet, 1, 21-2)

            # Generate Competitive Graphs (These will compare sites against other sites on basic revenue, traffic, discounts, etc.)
            # Create Data Tables (will be used to input the data into the visuals sheet for the corporation, tracking the cell inputs with python we can dynamically generate & gather graph data.)
            retail_quantity_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in retail_quantity_mix_data: 
                            retail_quantity_mix_data[location] = {}
                        retail_quantity_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Quantity'] - MONTHLY_STATS[location][washPkg]['Quantity']})
            

            retail_revenue_mix_data = {}
            total_retail_revenue_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_retail_revenue_mix_data[location] = COMBINED_SALES[location]['NET Sales'] - MONTHLY_SALES[location]['NET Sales']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in retail_revenue_mix_data: retail_revenue_mix_data[location] = {}
                        retail_revenue_mix_data[location].update({washPkg: (COMBINED_STATS[location][washPkg]['Quantity'] - MONTHLY_STATS[location][washPkg]['Quantity'])*COMBINED_STATS[location][washPkg]['Price']})
                    retail_revenue_mix_data[location].update({'Discounts': COMBINED_SALES[location]['Discounts']})

            monthly_quantity_mix_data = {}

            for location in MONTHLY_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in monthly_quantity_mix_data: 
                            monthly_quantity_mix_data[location] = {}
                        monthly_quantity_mix_data[location].update({washPkg: MONTHLY_STATS[location][washPkg]['Quantity']})
            

            monthly_revenue_mix_data = {}
            total_monthly_revenue_mix_data = {}

            for location in MONTHLY_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_monthly_revenue_mix_data[location] = MONTHLY_SALES[location]['NET Sales']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in monthly_revenue_mix_data: monthly_revenue_mix_data[location] = {}
                        monthly_revenue_mix_data[location].update({washPkg: MONTHLY_STATS[location][washPkg]['Amount']})
            

            combined_quantity_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    for washPkg in COMBINED_STATS[location]:
                        if location not in combined_quantity_mix_data: 
                            combined_quantity_mix_data[location] = {}
                        combined_quantity_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Quantity']})
            

            corporate_churn_rates_data = {}

            for location in ChurnTotal:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    if location not in corporate_churn_rates_data: 
                        corporate_churn_rates_data[location] = 0
                    corporate_churn_rates_data.update({location: ChurnTotal[location]})
            

            combined_revenue_mix_data = {}
            total_combined_revenue_mix_data = {}

            for location in COMBINED_STATS:
                if location not in [Corporation_Totals_Name, 'Query Server']:
                    total_combined_revenue_mix_data[location] = COMBINED_SALES[location]['NET Sales'] - COMBINED_SALES[location]['Discounts']
                    for washPkg in COMBINED_STATS[location]:
                        if location not in combined_revenue_mix_data: combined_revenue_mix_data[location] = {}
                        combined_revenue_mix_data[location].update({washPkg: COMBINED_STATS[location][washPkg]['Amount']})
            
            start_row = 75
            start_col = 1
            worksheet = wb[f'Sales Mix Charts (Corporation)']

            accounting_style_2 = NamedStyle(
                name='accounting_style_2', 
                number_format='$#,##0.00',  # or you could use 'Accounting'
                alignment=Alignment(horizontal='right')
            )
            percent_style_2 = NamedStyle(name='percent_style_2', number_format='0.00%')

            '''# Loop through each location and its corresponding wash packages
            for location, washPackages in retail_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row, column=start_col+1, value=location.strip("wash*u - "))
                
                # Increment the row to start adding wash packages for this location
                current_row = start_row + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    if start_col == 1:
                        worksheet.cell(row=current_row, column=start_col, value=washPackage)
                    worksheet.cell(row=current_row, column=start_col + 1, value=value)
                    current_row += 1  # Move down a row for the next wash package
                
                # Move to the next column to start adding the next location
                start_col += 1  # Shift 2 columns to the right for the next location'''
            # Loop through each location and its corresponding wash packages
            for location, washPackages in retail_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=75, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location

            createCorporateBarGraph(worksheet, "AB2", Reference(worksheet, min_col=2, min_row=75, max_col=5, max_row=start_row), Reference(worksheet, min_col=1, min_row=76, max_col=1, max_row=start_row), "Retail Wash Sales by Site (QTY)", y_axisTitle="# Washes Sold")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in retail_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_retail_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_retail_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown (%)", showPercents=True)


            start_row += 3  # Shift 2 rows down for the next location
            washLabelRow = start_row
# Loop through each location and its corresponding wash packages
            for location, washPackages in monthly_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            createCorporateBarGraph(worksheet, "AB37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Wash Rdmds by Site (QTY)", y_axisTitle="# Washes Rdmd")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in monthly_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Revenue Breakdown by Site ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_monthly_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_monthly_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN37", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Monthly Revenue Breakdown (%)", showPercents=True)




# Loop through each location and its corresponding wash packages
            start_row += 3  # Shift 2 rows down for the next location
            washLabelRow = start_row
            for location, washPackages in combined_quantity_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value)
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location

            createCorporateBarGraph(worksheet, "AB74", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Combined Wash Sales by Site (QTY)", y_axisTitle="# Washes Sold")

            
            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row

            for location, washPackages in combined_revenue_mix_data.items():
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                # Loop through each wash package and add it to the sheet
                for washPackage, value in washPackages.items():
                    worksheet.cell(row=washLabelRow, column=current_col, value=washPackage)
                    worksheet.cell(row=start_row + 1, column=current_col, value=value).style = accounting_style_2
                    current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            #createCorporateBarGraph(visualWorksheet, "AM2", Reference(visualWorksheet, min_col=2, min_row=washLabelRow, max_col=6, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Retail Revenue Breakdown ($)", useDataTitle=True)

            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in total_combined_revenue_mix_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=total_combined_revenue_mix_data[location]).style = accounting_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createCorporatePieChart(visualWorksheet, "AN74", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=5, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Combined Revenue Breakdown ($)", showPercents=True)


            start_row += 1  # Shift 2 rows down for the next location
            washLabelRow = start_row
            
            for location in corporate_churn_rates_data:
                # Write location name to the sheet
                worksheet.cell(row=start_row + 1, column=start_col, value=location.strip("wash*u - "))
                
                
                # Increment the column to start adding wash packages for this location
                current_col = start_col + 1
                
                worksheet.cell(row=start_row + 1, column=current_col, value=corporate_churn_rates_data[location]).style = percent_style_2
                current_col += 1  # Move to the next column for the next wash package
                
                # Move to the next row to start adding the next location
                start_row += 1  # Shift 2 rows down for the next location
            
            createBarGraph(visualWorksheet, "P71", Reference(visualWorksheet, min_col=2, min_row=washLabelRow+1, max_col=2, max_row=start_row), Reference(visualWorksheet, min_col=1, min_row=washLabelRow+1, max_col=1, max_row=start_row), "Churn Rate by Location (%)")




    #--> Modify the template and populate all blocks with site data
    wb = openpyxl.load_workbook(templateFilePath)
    ws = wb['Sales Mix by Location']

    start_col = 1
    #--> Assign a report date to the report.
    ws.cell(row=1, column=2, value=complete_date_str)
    locations = list(MONTHLY_STATS.keys())
    locations.remove("Corporation Totals")
    if "Query Server" in locations:
        locations.remove("Query Server")
    for location in locations:
        print(f"Creating {location} Stats Block...")
        copy_cells(ws, 3, 30, 1, 7, 3, start_col) #copy_cells(ws, 1, 28, 1, 6, 1, start_col)
        populate_salesmix_worksheet(ws, 3, start_col, location)
        start_col += 8
    
    copy_cells(ws, 3, 30, 1, 7, 32, 1)
    populate_salesmix_worksheet(ws, 32, 1, Corporation_Totals_Name)

    if not "Query Server" in blacklisted_sites:
        print("Creating Query Server Stats Block...")
        copy_cells(ws, 3, 30, 1, 7, 32, 9)
        populate_salesmix_worksheet(ws, 32, 9, "Query Server")

    
        
    #--> Discounts Worksheet Stuff
    # Define styles
    int_style = NamedStyle(name='integer_style', number_format='#,##0')
    accounting_style = NamedStyle(
        name='accounting_style', 
        number_format='#,##0.00',  # or you could use 'Accounting'
        alignment=Alignment(horizontal='right')
    )
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    # Modify the populate_discountmix_worksheet function to include formatting
    def populate_discountmix_worksheet(worksheet, start_row, start_col, location, discount_data):
        worksheet.cell(row=start_row, column=start_col, value=location)
        
        # Adding headers
        headers = ["Discount Name", "Quantity Used", "Price*", "Total Discount Dollars", "Discount Distribution (Qty)", "Discount Distribution ($)"]
        for j, header in enumerate(headers):
            worksheet.cell(row=start_row + 2, column=start_col + j, value=header)
            
        total_quantity = 0
        total_amount = 0

        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            total_quantity += discount_info['Quantity'] if discount_info['Quantity'] is not None else 0
            total_amount += discount_info['Amount'] if discount_info['Amount'] is not None else 0
        
        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            if discount_name is not None:

                # Add discount name
                worksheet.cell(row=start_row+3+i, column=start_col, value=discount_name)
                
                # Add and format Quantity
                cell_quantity = worksheet.cell(row=start_row+3+i, column=start_col+1, value=discount_info['Quantity'])
                cell_quantity.style = int_style
                
                # Add and format Price
                cell_price = worksheet.cell(row=start_row+3+i, column=start_col+2, value=discount_info['Price'])
                cell_price.style = accounting_style
                
                # Add and format Amount
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+3, value=discount_info['Amount'])
                cell_amount.style = accounting_style

                # Add and format Distribution by Quantity
                distribution_value = discount_info['Quantity'] / total_quantity if total_quantity != 0 else 0
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+4, value=distribution_value)
                cell_amount.style = percent_style  # Apply the percentage style

                # Add and format Distribution by Amount
                distribution_value = discount_info['Amount'] / total_amount if total_amount != 0 else 0
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+5, value=distribution_value)
                cell_amount.style = percent_style  # Apply the percentage style
                
                
        '''
        # Add and format "Total"
        total_label_cell = worksheet.cell(row=start_row+4+i, column=start_col, value="Total")
        total_label_cell.font = Font(bold=True)

        # Add and format Total Quantity
        cell_total_quantity = worksheet.cell(row=start_row+4+i, column=start_col+1, value=total_quantity)
        cell_total_quantity.style = int_style  # If you have defined int_style elsewhere
        cell_total_quantity.font = Font(bold=True)

        # Add and format Total Amount
        cell_total_amount = worksheet.cell(row=start_row+4+i, column=start_col+3, value="${:.2f}".format(total_amount))
        cell_total_amount.style = accounting_style  # If you have defined accounting_style elsewhere
        cell_total_amount.font = Font(bold=True)
        '''
        
        # Add a Price disclaimer explaining that some prices are averages
        if location == Corporation_Totals_Name:
            disclaim1 = worksheet.cell(row=start_row+4+i+2, column=start_col, value="NOTES")
            disclaim2 = worksheet.cell(row=start_row+4+i+3, column=start_col, value="Price*: Some 'Open Price' items such as 'Misc. Wash Discount' show the average amount that was paid out over the period under that discount item.")
            disclaim3 = worksheet.cell(row=start_row+4+i+4, column=start_col, value="Discounts that are not applied (are missing on the GSR for the location while parsing; thus quantity is 0) will not be shown on the report for that location.")
            disclaim4 = worksheet.cell(row=start_row+4+i+5, column=start_col, value="Due to API limitations with Excel, a proper Total Row for the table cannot be 'enabled' by default. Therefore, it is recommended that you (the end user) enable the total row by clicking anywhere in the table, click table design in the ribbon bar and then check 'Total Row' in the Table Style Options section.")
            disclaim1.font = Font(bold=True)
            disclaim2.font = Font(bold=True)
            disclaim3.font = Font(bold=True)
            disclaim4.font = Font(bold=True)

        tableName = re.sub(r'\W+', '', location.replace(" ", ""))
        # Adding table and table banding
        tab = Table(displayName=tableName, ref=f"{worksheet.cell(row=start_row + 2, column=start_col).coordinate}:{worksheet.cell(row=start_row + 4 + i, column=start_col + 5).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium23", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    # Read the input CSV to a DataFrame
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    # Aggregate the data
    discount_data = {}
    miscDisctAverages = {}
    discountTotals = {}
    discount_data[Corporation_Totals_Name] = {}
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        discount_name = row['Item Name']
        quantity = row['Quantity']
        price = row['Price']
        amount = row['Amount']

        if (category in Discount_Categories and 
            site not in blacklisted_sites and 
            pd.notna(discount_name) and 
            pd.notna(quantity) and 
            pd.notna(price) and 
            pd.notna(amount)):
            
            if site not in discount_data:
                discount_data[site] = {}
                
            if discount_name not in discount_data[site]:
                discount_data[site][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            if discount_name not in discount_data[Corporation_Totals_Name]:
                discount_data[Corporation_Totals_Name][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            discount_data[site][discount_name]['Quantity'] += quantity
            discount_data[Corporation_Totals_Name][discount_name]['Quantity'] += quantity
            if discount_name not in miscDisctAverages:
                miscDisctAverages[discount_name] = []
            miscDisctAverages[discount_name].append(price if pd.notna(price) else 0)
            discount_data[site][discount_name]['Amount'] += amount
            discount_data[Corporation_Totals_Name][discount_name]['Amount'] += amount
    
    # Average the prices
    # Used for Misc. discounts to see the average amount of money paid out to customers
    locations = list(discount_data.keys())

    for location in locations:
        for discount in discount_data[location]:
            try:
                discount_data[location][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
                discount_data[Corporation_Totals_Name][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
            except:
                discount_data[location][discount]['Price'] = 0

    # Load the template workbook
    ws = wb['Discount Mix by Location']
    # Initial starting column for the first location
    start_col = 1

    # Populate the worksheet
    for location, data in discount_data.items():
        copy_cells(ws, 1, 3, 1, 6, 1, start_col)
        populate_discountmix_worksheet(ws, 1, start_col, location, data)
        start_col += 7
        
    wb._sheets.remove(wb['Sales Mix by Location (Visual)'])


    wbName = f"Sales Mix - {short_date_str}"
    return wb, wbName

    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################
    #########################################################################################################################################################################

def appendToAWPWorkbook(gsrFilePath, templateFilePath, fileNameForParser, AWP_Template, overrideInputValidation=None):
    #--> Kind of a lazy appending function that takes a lot of the data from the Sales Mix report code and uses it here.
    #--> This should help standardize the data we view at WashU making more reports use the same numbers from the same data source(s). 
    #--> This is a RUNNING WORKSHEET, which means that each month a new version is created when the correct GSR and setting is turned on in the portal.

    #--> Create Dictionaries for looking up item names.
    # TODO: This needs to be attached to a database and allow insertions through the manager portal or a form.
    global monthsInReport
    monthsInReport = 1
    MONTHLY_STATS = {}
    COMBINED_STATS = {}
    COMBINED_SALES = {}
    MONTHLY_SALES = {}

    ARM_PKG_Names = {
        "Express Wash": {
            "New Mnthly Exp Rdmd",
            "New Monthly Expr Rdm",
            #"3 Mo Express Rdmd",
            #"CLUB-express Rdmd",
            #"COMP-CLUB-xprs Rdmd",
        },
        "Clean Wash": {
            #"3 Mo Clean Rdmd",
            #"CLUB-clean Rdmd",
            "Monthly Cln '21 Rdmd",
            #"COMP-CLUB-clean Rdmd",
            "W-Mon Clean V Rdmd",
        },
        "Protect Wash": {
            #"3 Mo Protect Rdmd",
            #"City Fire Club Rdmd",
            #"CLUB-protect Rdmd",
            #"COMP-CLUB-prot Rdmd",
            "Mnthly Prot 6mo Rdmd",
            "New Mnthy Prot Rdmd",
            "New Mnthly Prot Rdmd",
        },
        "UShine Wash": {
            "Monthly UShine Rdmd",
            #"3 Mo UShine Rdmd",
            #"CLUB-ushine Rdmd",
            #"COMP-CLUB-ushin Rdmd",
            #"Free Week UShine Rdm",
        },
    }
    Monthly_Total_Categories = [
        # Member revenue is calculated by adding amounts from these categories.
        "ARM Plans Sold", 
        "ARM Plans Recharged", 
        #"Club Plans Sold", 
        "ARM Plans Terminated"
    ]
    ARM_Sold_Names = {
        "Express Wash": {
            "New Mnthly Express",
            "Mnth Exp 9.95 (BOGO)", # Added 10/16/23
            "New Mnthly Exp 9.95" # Added 10/16/23
        },
        "Clean Wash": {
            "Mnth Cln 9.95 (BOGO)",
            "Monthly Cln '21 9.95",
            "Monthly Cln '21 Sld",
            "3 Mo CleanClubSldTS",
        },
        "Protect Wash": {
            "Mnthly Protect Promo",
            "New Mnthly Protect",
            "City Fire Club Sld",
            "3 Mo Protect ClubSld"
        },
        "UShine Wash": {
            "Mnth USh 9.95 (BOGO)",
            "Monthly UShine Promo",
            "Monthly UShine Sld",
            "Mnth USh 9.95 (BOGO)" # Added 10/16/23
        },
    }


    #--> Utility Functions
    def find_parent(json_obj, target_str, start_parent=None, current_parent=None):
        """
        Searches for a string in a nested JSON object (Python dict) starting from a specified parent,
        and returns the parent key under which the string resides.
        
        Parameters:
        - json_obj (dict): The JSON object to search through.
        - target_str (str): The string to look for.
        - start_parent (str): The parent key from where to start the search.
        - current_parent: The current parent key, used for recursion.
        
        Returns:
        - The parent key under which the string resides, or None if the string is not found.
        """
        if start_parent and current_parent is None:
            json_obj = json_obj.get(start_parent, {})
            
        for key, value in json_obj.items():
            if key == target_str:
                return current_parent
            if isinstance(value, dict):
                result = find_parent(value, target_str, None, key)
                if result:
                    return result
            elif isinstance(value, (list, set)):
                if target_str in value:
                    return key
        return None
    
    
    #########################################################################################################################################################################
    ############################################################################## GSR PARSING ##############################################################################
    #########################################################################################################################################################################
    #--> Gather Start/End Dates for the report

    # Read the CSV file to extract the start date and end date
    df = pd.read_csv(gsrFilePath, nrows=2)  # Reading only the first 2 rows
    global GSR_YEAR
    global GSR_MONTH
    GSR_YEAR = 0
    GSR_MONTH = 0
    oneCalendarMonthGSR = False
    monthsInReport = None

    def is_valid_date(date_str, date_format='%Y-%m-%d'):
        try:
            datetime.strptime(date_str, date_format)
            return True
        except ValueError:
            return False

    def extract_dates_from_filepath():
        # Initialize variables
        global GSR_MONTH
        global GSR_YEAR
        global monthsInReport
        global oneCalendarMonthGSR
        
        # Use regular expression to find dates in the filename
        match = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})", fileNameForParser)
        
        if match:
            start_date_str, end_date_str = match.groups()
            
            # Check if dates are valid
            if not is_valid_date(start_date_str) or not is_valid_date(end_date_str):
                print("Invalid date(s) in filename. Skipping...")
                return None, None, None
            
            # Convert the extracted date strings to datetime objects
            start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d')

            start_date = start_date_dt.date()
            end_date = end_date_dt.date()

            GSR_YEAR = start_date.year
            GSR_MONTH = start_date.month

            # Calculate the difference in days between the two dates
            delta = (end_date_dt - start_date_dt).days

            _, last_day = calendar.monthrange(start_date.year, start_date.month)

            # Directly compare start_date and end_date to determine if the report spans one calendar month
            oneCalendarMonthGSR = (start_date.year == end_date.year) and (
                (end_date.month == start_date.month and end_date.day == last_day) or
                (end_date.month == start_date.month + 1 and end_date.day == start_date.day - 1)
            )
                
            # Calculate the number of months (keeping your original logic here)
            monthsInReport = (delta / 30.44) if (delta / 30.44) > 1 else 1
            
            if not oneCalendarMonthGSR:
                print("Uploaded GSR is not one calendar month.. script will error.")
                
            return start_date, end_date, oneCalendarMonthGSR
        else:
            return None, None, None

    # Example usage
    start_date, end_date, oneCalendarMonthGSR = extract_dates_from_filepath()
    print(start_date, end_date)
    # Format the date strings
    short_date_str, complete_date_str = str(start_date) + ' - ' + str(end_date), str(start_date) + ' - ' + str(end_date)

    if oneCalendarMonthGSR == False and overrideInputValidation == None:
        return "error501", f"The GSR you submitted spanning date range: {complete_date_str} is not one calendar month and therefore will not be added to the provided template. Please download a calendar month GSR and upload that along with your specified databook template."
        
    #--> Prepare dictionaries for data entry
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        print("Calculating for", site)
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']

        if site not in blacklisted_sites:
            if site not in MONTHLY_STATS:
                MONTHLY_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if site not in COMBINED_STATS:
                COMBINED_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            if Corporation_Totals_Name not in MONTHLY_STATS:
                MONTHLY_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0, 'Gross New Members': 0}}
            if Corporation_Totals_Name not in COMBINED_STATS:
                COMBINED_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            
            if site not in MONTHLY_SALES:
                MONTHLY_SALES[site] = {"NET Sales": 0}
            if site not in COMBINED_SALES:
                COMBINED_SALES[site] = {"Discounts": 0, "NET Sales": 0}
            if Corporation_Totals_Name not in MONTHLY_SALES:
                MONTHLY_SALES[Corporation_Totals_Name] = {"NET Sales": 0}
            if Corporation_Totals_Name not in COMBINED_SALES:
                COMBINED_SALES[Corporation_Totals_Name] = {"Discounts": 0, "NET Sales": 0}

    #--> Gather TOTAL sales & discounts details
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            COMBINED_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
            COMBINED_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            if category in Discount_Categories:
                COMBINED_SALES[site]['Discounts'] += amount if pd.isna(amount) != True else 0
                COMBINED_SALES[Corporation_Totals_Name]['Discounts'] += amount if pd.isna(amount) != True else 0

            # Handle the Query Server
            if site == 'Query Server':
                if find_parent(Website_PKG_Names, item_name, start_parent="Monthly") != None or category in Monthly_Total_Categories:
                    MONTHLY_SALES['Query Server']['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0
            # Any other locations
            else: 
                if category in Monthly_Total_Categories:
                    MONTHLY_SALES[site]['NET Sales'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if pd.isna(amount) != True else 0

            

    #--> Gather Sales Mix
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    mask = ~df['Site'].isin(blacklisted_sites)
    df = df[mask]

    MonthlyAmounts = {}
    CombinedAmounts = {}
    AccurateMemberCountAdjuster = {}
    ChurnDataDictionary = {}
    ChurnTotal = {}
    totalCorporateDiscont = 0

    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        count = row['Count']
        price = row['Price']
        quantity = row['Quantity']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items and item_name not in blacklisted_items:

            if ChurnDataDictionary.get(site) == None:
                ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            if ChurnDataDictionary.get(Corporation_Totals_Name) == None:
                ChurnDataDictionary[Corporation_Totals_Name] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
            
            if AccurateMemberCountAdjuster.get(site) == None:
                AccurateMemberCountAdjuster[site] = 0
            if item_name in ['WEB Discontinue ARM', 'Discontinue ARM Plan']:
                if AccurateMemberCountAdjuster.get(site) != None:
                    AccurateMemberCountAdjuster[site] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    totalCorporateDiscont += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    ChurnDataDictionary[site]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                    ChurnDataDictionary[Corporation_Totals_Name]['Discontinued'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 

            washPkg = find_parent(ARM_Recharge_Names, item_name) # Member Count = ARM Recharges
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Recharged'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 

            washPkg = find_parent(Website_PKG_Names, item_name, "Monthly") # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 

            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld | Gross New Members = ARM Sold
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[site][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
            
            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                ChurnDataDictionary[site]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Sold'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 

            washPkg = find_parent(ARM_Termination_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld - ARM Terminations
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Subtract the terminated quantity from the estimated count. We divide by the months in the report to correct for multi-month GSRs
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[site][washPkg]['Estimated Member Count'])
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] -= round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] = abs(MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'])
                ChurnDataDictionary[site]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Terminated'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
            
            if item_name == "Switch ARM Plan":
                ChurnDataDictionary[site]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 
                ChurnDataDictionary[Corporation_Totals_Name]['Switched'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 

            # Handle the Query Server
            if site == 'Query Server':
                # We're only going to show the amount of passes sold rather than the redemption for the
                # Query Server because you cannot redeem passes online.
                washPkg = find_parent(Website_PKG_Names, item_name, "Monthly")
                if washPkg != None:
                    MONTHLY_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0 # Gross New Members
    
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Gross New Members'] += round(quantity/monthsInReport) if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0)
                
                washPkg = find_parent(Website_PKG_Names, item_name, "Retail")
                if washPkg != None:
                    COMBINED_STATS['Query Server'][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS['Query Server'][washPkg]['Price'] += price if pd.isna(price) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        CombinedAmounts[washPkg] = []
                    CombinedAmounts[washPkg].append(price if pd.isna(price) != True else 0)

            # Any other locations
            else: 
                washPkg = find_parent(ARM_PKG_Names, item_name)
                # Calculate MONTHLY Stats using the redemption items
                if category in ["ARM Plans Redeemed", "Club Plans Redeemed"] and washPkg is not None:
                    MONTHLY_STATS[site][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[site][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[site][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    MONTHLY_STATS[site][washPkg]['Price'] += abs(price) if pd.isna(price) != True else 0

                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if pd.isna(count) != True else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if pd.isna(price) != True else 0) 
                elif category in ['Basic Washes'] and item_name is not None and item_name in COMBINED_STATS[site]:
                    COMBINED_STATS[site][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[site][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[site][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    COMBINED_STATS[site][item_name]['Price'] += amount/quantity if pd.isna(quantity) != True else 0

                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'] += quantity if pd.isna(quantity) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Count'] += count if pd.isna(count) != True else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Amount'] += amount if pd.isna(amount) != True else 0
                    if not washPkg in CombinedAmounts:
                        CombinedAmounts[item_name] = []
                    CombinedAmounts[item_name].append(price if pd.isna(price) != True else 0)

    for washPkg in ARM_Sold_Names:
        COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount']/COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount']/MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity']
        #COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(CombinedAmounts[washPkg])/len(CombinedAmounts[washPkg])
        #MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(MonthlyAmounts[washPkg])/len(MonthlyAmounts[washPkg])
    AccurateMemberCountAdjuster[Corporation_Totals_Name] = totalCorporateDiscont

    for location, stats in ChurnDataDictionary.items():
        # ChurnDataDictionary[site] = {"Terminated":0, "Discontinued":0, "Recharged":0, "Sold":0, "Switched":0}
        ChurnTotal[location] = (ChurnDataDictionary[location]["Terminated"]+ChurnDataDictionary[location]["Discontinued"]) / (ChurnDataDictionary[location]["Recharged"]+ChurnDataDictionary[location]["Sold"]+ChurnDataDictionary[location]["Discontinued"]-ChurnDataDictionary[location]["Switched"])

    print(MONTHLY_STATS)
    print("")
    print(MONTHLY_SALES)
    #########################################################################################################################################################################
    ############################################################################## XCL LOADING ##############################################################################
    #########################################################################################################################################################################
    
    # Load the workbook
    wb = openpyxl.load_workbook(AWP_Template)

    # Select the sheet you want to work with
    ws = wb["Wash Counts"]

    # Initialize an empty dictionary to store the values and their positions
    value_position_dict = {}
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    total_cars_washed = 0


    # Loop through the first row
    for cell in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for individual_cell in cell:
            value = individual_cell.value
            position = individual_cell.coordinate
            if value:
                value_position_dict[value] = position
                if value == "Total Cars Washed":
                    break
        if "Total Cars Washed" in value_position_dict:
            break

    # Function to convert cell coordinates to row, col tuple
    def coordinate_to_tuple(coordinate):
        col_str = ''
        row_str = ''
        for char in coordinate:
            if char.isalpha():
                col_str += char
            else:
                row_str += char
        row = int(row_str)
        col = 0
        for i, char in enumerate(reversed(col_str)):
            col += (ord(char.upper()) - ord('A') + 1) * (26 ** i)
        return row, col

    # Initialize starting cell based on the first value in the dictionary
    starting_cell = list(value_position_dict.values())[0]

    # Convert coordinate to row and column tuple
    row, col = coordinate_to_tuple(starting_cell)

    # Move down 3 rows to row 4
    row += 3

    # Find the first blank cell for the first label
    while ws.cell(row=row, column=col).value is not None:
        row += 1

    # Insert a new row for the new month
    ws.insert_rows(row + 1)

    # Create a NamedStyle object for reusability
    copy_style = NamedStyle(name="copy_style")

    # Loop through all labels in value_position_dict
    for currentLabel, starting_cell in value_position_dict.items():
        if currentLabel not in COMBINED_STATS or not COMBINED_STATS[currentLabel]:
            if currentLabel == 'Total Cars Washed':
                # Get the formula from the cell above
                above_formula = ws.cell(row=row - 1, column=col).value
                
                if isinstance(above_formula, str) and above_formula.startswith("="):
                    # Update the row numbers in the formula
                    updated_formula = above_formula.replace(str(row - 1), str(row))
                    
                    # Set the updated formula to the current "Total Cars Washed" cell
                    ws.cell(row=row, column=col).value = updated_formula
            col += 6
            continue
            # Copy the formatting for all cells in the new row from the row above
        for col_num in range(1, ws.max_column + 1):
            new_cell = ws.cell(row=row + 1, column=col_num)
            above_cell = ws.cell(row=row, column=col_num)
            
            if above_cell.has_style:
                new_cell._style = copy(above_cell._style)

        # Insert the data here for Month, Car Count, and NET Sales
        left_cell_value = ws.cell(row=row, column=col - 3).value
        if len(left_cell_value.split()) > 1:
            if left_cell_value.split()[1] == 'Jan':
                next_month_index = (months.index(left_cell_value.split()[1]) + 1) % 12
            else:
                print("Invalid left cell value for date. Expecting 2 parts, str|num got unexpected.")
                break
        else:
            next_month_index = (months.index(left_cell_value) + 1) % 12
                
        next_month = months[next_month_index]
        if next_month == 'Jan':
            newYear = ws.cell(row=row-9, column=col - 3).value.split()[0]
            ws.cell(row=row + 1, column=col - 3).value = f"{int(newYear)+1} Jan"
        else:
            ws.cell(row=row + 1, column=col - 3).value = next_month

        total_cars_washed += sum(washItemData.get('Quantity', 0) for washItemData in COMBINED_STATS[currentLabel].values())
        car_count = sum(washItemData.get('Quantity', 0) for washItemData in COMBINED_STATS[currentLabel].values())
        ws.cell(row=row, column=col).value = car_count

        NET_Wash_Sales = COMBINED_SALES[currentLabel]['NET Sales']# - MONTHLY_SALES[currentLabel]['NET Sales']
        ws.cell(row=row, column=col + 1).value = NET_Wash_Sales

        col += 6  # Move over 6 columns for the next label

    #--> Worksheet: Membership Data
    ws2 = wb['Membership Data']

    # Initialize variable to store the coordinate of the first blank cell in row 3
    first_blank_in_row3 = None

    # Loop through the cells in row 3 starting from column F (column index 6)
    for cell in ws2.iter_rows(min_row=3, max_row=3, min_col=6, max_col=ws2.max_column):
        for individual_cell in cell:
            if individual_cell.value is None:
                first_blank_in_row3 = individual_cell.coordinate
                break  # Break the inner loop
        if first_blank_in_row3:
            break  # Break the outer loop if we found the first blank cell

    print(f"First Blank Cell in Row 3: {first_blank_in_row3}")

    # Initialize starting column (F) and row (3)
    col = 6  # Column F
    row = 3

    # Parse row 3 until we find a blank cell, starting from column F
    while ws2.cell(row=row, column=col).value is not None:
        col += 1

    # Capture the cell above the blank one for the month and year
    above_cell_value = ws2.cell(row=row - 1, column=col).value

    # Extract the month and year
    month, year = above_cell_value.split()
    year = int(year)

    if int("20"+str(year)) != int(GSR_YEAR) and overrideInputValidation == None:
        return "error502", f"There was an error loading the data into the workbook. It seems that the GSR you submitted is from {GSR_YEAR}, while the databook template you uploaded is ready to have {year} data loaded into it. "
    if months.index(month) != GSR_MONTH-1 and overrideInputValidation == None:
        return "error503", f"There was an error loading the data into the workbook. It seems that the GSR you submitted is from {months[GSR_MONTH-1]}, while the databook template you uploaded is ready to have {month} data loaded into it. "

    # Define the list of months
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Find the next month and year
    next_month_index = (months.index(month) + 1) % 12
    next_month = months[next_month_index]
    if next_month == 'Jan':
        year += 1

    # Insert a new column ahead of the current column
    ws2.insert_cols(col + 1)

    # Copy the formatting and values/formulas for all cells in the new column from the column to the left
    for row_num in range(1, ws2.max_row + 1):

        new_cell2 = ws2.cell(row=row_num, column=col)
        left_cell2 = ws2.cell(row=row_num, column=col - 1)
        
        if left_cell2.has_style:
            new_cell2._style = copy(left_cell2._style)
        
        new_cell = ws2.cell(row=row_num, column=col + 1)
        left_cell = ws2.cell(row=row_num, column=col)
        
        if left_cell.has_style:
            new_cell._style = copy(left_cell._style)

    # Set the value for the new month and year in the new column
    ws2.cell(row=row - 1, column=col + 1).value = f"{next_month} {str(year)[-2:]}"

    #--> Insert Data

    def update_formula(formula, col_shift):
        
        def replacer(match):
            col_letter = match.group(1)
            col_idx = column_index_from_string(col_letter)
            new_col_idx = col_idx + col_shift
            new_col_letter = get_column_letter(new_col_idx)
            return new_col_letter + match.group(2)
        
        return re.sub(r'([A-Z]+)([0-9]+)', replacer, formula)

    # Loop for Estimated Member Count
    while ws2.cell(row=row, column=5).value != 'Combined':
        store = ws2.cell(row=row, column=5).value
        if store:  # Check if store is not None or empty
            prefixedStore = f"wash*u - {store}" if store != "Admin" else "Query Server"
            if prefixedStore in MONTHLY_STATS:
                total_estimated_count = sum(wash_data.get("Estimated Member Count", 0) for wash_data in MONTHLY_STATS[prefixedStore].values()) - AccurateMemberCountAdjuster[prefixedStore]
                ws2.cell(row=row, column=col).value = total_estimated_count
            
        row += 1
    
    # Copy and modify formula for the new "Combined" cell (Estimated Member Count)
    formula_to_copy = ws2.cell(row=row, column=col-1).value
    print(formula_to_copy)
    if formula_to_copy.startswith('='):
        new_formula = update_formula(formula_to_copy, 1)
        ws2.cell(row=row, column=col).value = new_formula
        

    # Skip 2 rows to go to the NET Sales block
    row += 2

    # Loop for NET Sales
    while ws2.cell(row=row, column=5).value != 'Combined':
        store = ws2.cell(row=row, column=5).value
        if store:  # Check if store is not None or empty
            prefixedStore = f"wash*u - {store}" if store != "Admin" else "Query Server"
            if prefixedStore in MONTHLY_SALES:
                net_sales = MONTHLY_SALES[prefixedStore].get("NET Sales", 0)
                ws2.cell(row=row, column=col).value = net_sales
        row += 1

    # Copy and modify formula for the new "Combined" cell (Estimated Member Count)
    formula_to_copy = ws2.cell(row=row, column=col-1).value
    if formula_to_copy.startswith('='):
        new_formula = update_formula(formula_to_copy, 1)
        ws2.cell(row=row, column=col).value = new_formula

    # Skip 2 rows to go to the Quantity block
    row += 2

    # Loop for Quantity
    while ws2.cell(row=row, column=5).value != 'Combined':
        store = ws2.cell(row=row, column=5).value
        if store:  # Check if store is not None or empty
            prefixedStore = f"wash*u - {store}"
            if prefixedStore in MONTHLY_STATS:
                total_quantity = sum(wash_data.get("Quantity", 0) for wash_data in MONTHLY_STATS[prefixedStore].values())
                ws2.cell(row=row, column=col).value = total_quantity
        row += 1

        # Copy and modify formula for the new "Combined" cell (Estimated Member Count)
    formula_to_copy = ws2.cell(row=row, column=col-1).value
    print(formula_to_copy)
    if formula_to_copy.startswith('='):
        new_formula = update_formula(formula_to_copy, 1)
        ws2.cell(row=row, column=col).value = new_formula


    for row_num in range(row + 1, ws2.max_row + 1):
        new_cell = ws2.cell(row=row_num, column=col)
        left_cell = ws2.cell(row=row_num, column=col - 1)
        
        if left_cell.value and "=" in str(left_cell.value):  # Check if it's a formula
            new_formula = update_formula(left_cell.value, 1)
            new_cell.value = new_formula

        if left_cell.has_style:
            new_cell._style = copy(left_cell._style)


    # Save the workbook
    wb.save("NCW_Manager_Portal/your_file.xlsx")
    
    wbName = f"Membership Retail and AWP Databook - {short_date_str}"
    return wb, wbName
