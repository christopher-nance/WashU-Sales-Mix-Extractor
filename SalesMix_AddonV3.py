# Built by Christopher Nance for WashU Car Wash
# Version 3.0
# Sales Mix Report Generator

# Dependencies:
# > Python 3.10
# > Pandas
# > Power Automate
# > Dropbox
# > SiteWatch Cloud (SiteWatch Version 26+)
# > Microsoft Azure Cloud Functions


import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formula.translate import Translator
from copy import copy
from datetime import datetime
import re
import json



def createSalesMixSheet(gsrFilePath, trendsFilePath=None, templateFilePath='/generated_reports/Sales_Mix_Template.xlsx'):
    #--> Create Dictionaries for looking up item names.
    # TODO: This needs to be attached to a database and allow insertions through the manager portal or a form.
    Corporation_Totals_Name = 'Corporation Totals'
    ARM_Sold_Names = {
        "Express Wash": {
            "New Mnthly Express",
        },
        "Clean Wash": {
            "Mnth Cln 9.95 (BOGO)",
            "Monthly Cln '21 9.95",
            "Monthly Cln '21 Sld",
        },
        "Protect Wash": {
            "Mnthly Protect Promo",
            "New Mnthly Protect",
        },
        "UShine Wash": {
            "Mnth USh 9.95 (BOGO)",
            "Monthly UShine Promo",
            "Monthly UShine Sld",
        },
    }

    ARM_Recharge_Names = {
        "Express Wash": {
            "New Mnthly Exp Rchg",
            "New Monthly Expr Rcg",
        },
        "Clean Wash": {
            "Monthly Cln '21 Rchg",
        },
        "Protect Wash": {
            "New Mnthly Prot Rchg",
            "Mnthly Prot 6mo Rchg"
        },
        "UShine Wash": {
            "Monthly UShine Rchg"
        },
    }

    ARM_Termination_Names = {
        "Express Wash": {
            "New Mnthly Exp NoRfn",
            "New Mnthly Exp Rfnd",
            "New Monthly Exp Rfnd",
            "New Monthly ExpNoRfn",
        },
        "Clean Wash": {
            "Monthly Cln '21 Rfnd",
            "MonthlyCln'21 NoRfnd",
        },
        "Protect Wash": {
            "Mnthly Prot 6m NRfnd",
            "Mnthly Prot 6mo Rfnd",
            "New Mnthly Pro NoRfn",
            "New Mnthly Prot Rfnd",
        },
        "UShine Wash": {
            "Monthly UShine NRfnd",
            "Monthly UShine Rfnd",
        },
    }

    ARM_PKG_Names = {
        "Express Wash": {
            "New Mnthly Exp Rdmd",
            "New Monthly Expr Rdm",
            "3 Mo Express Rdmd",
            "CLUB-express Rdmd",
            "COMP-CLUB-xprs Rdmd",
        },
        "Clean Wash": {
            "3 Mo Clean Rdmd",
            "CLUB-clean Rdmd",
            "Monthly Cln '21 Rdmd",
            "COMP-CLUB-clean Rdmd",
        },
        "Protect Wash": {
            "3 Mo Protect Rdmd",
            "City Fire Club Rdmd",
            "CLUB-protect Rdmd",
            "COMP-CLUB-prot Rdmd",
            "Mnthly Prot 6mo Rdmd",
            "New Mnthy Prot Rdmd",
            "New Mnthly Prot Rdmd",
        },
        "UShine Wash": {
            "Monthly UShine Rdmd",
            "3 Mo UShine Rdmd",
            "CLUB-ushine Rdmd",
            "COMP-CLUB-ushin Rdmd",
            "Free Week UShine Rdm",
        },
    }
    Website_PKG_Names = {
        # The packages sold on the website differ in name and category than the other standard location packages. 
        # For instance, all website items (including ARM Sellers & Giftcard Sellers) are lumped into one "Website Sold" category
        "Retail": {
            "Express Wash": {
                
            },
            "Clean Wash": {
                "W-1-clean wash",
            },
            "Protect Wash": {
                "W-1-protect wash",
            },
            "UShine Wash": {
                "W-1-UShine wash",
            },
        },

        "Monthly": {
            "Express Wash": {
                "W-New MonthlyExprSld",
            },
            "Clean Wash": {
                "W-Mnthly Cln '21 Sld",
            },
            "Protect Wash": {
                "W-Unl. Prot 9.95 Sld",
                "W-Unl. protect Sld",
            },
            "UShine Wash": {
                "W-MonthlyUShine Sld",
                "W-UShine 9.95 Sld",
            },
        }
    }
    blacklisted_items = [
        # Copied directly from the GSR CSV, hence the duplicate values.
        "Deposits",
        "XPT Cash Add/Remove",
        "XPT Cash Add/Remove",
        "XPT Cash Add/Remove",
        "XPT Cash Add/Remove",
        "XPT Cash Add/Remove",
        "House Acct Related",
        "House Acct Related",
        "House Acct Related",
        "House Acct Related",
        "House Acct Related",
        "House Acct Related",
        "Employee Accounts",
        "Employee Accounts",
        "Cash",
        "XPT Cash Over/Short",
        "XPT Chg Over/Short",
        "Credit Card",
        "Credit Card",
        "Credit Card",
        "Other Tenders",
        "Other Tenders",
        "House Accounts",
        "XPT Balancing",
        "Picture Mismatch",
        "Employees"
        #"Website Sold"
    ]
    blacklisted_sites = [
        "Hub Office",
        "wash*u - Centennial",
        "Zwash*u - Centennial Old",
    ]
    Monthly_Total_Categories = [
        # Member revenue is calculated by adding amounts from these categories.
        "ARM Plans Sold", 
        "ARM Plans Recharged", 
        "Club Plans Sold", 
        "ARM Plans Terminated"
    ]
    ARM_Member_Categories = [
        "ARM Plans Sold", 
        "ARM Plans Recharged", 
        "ARM Plans Terminated"
    ]
    Discount_Categories = [
        "Website Discounts", 
        "Wash Discounts", 
        "Wash LPM Discounts"
    ]
    MONTHLY_STATS = {}
    COMBINED_STATS = {}
    COMBINED_SALES = {}
    MONTHLY_SALES = {}

    #--> Utility Functions
    def find_parent(json_obj, target_str, start_parent=None, current_parent=None):
        """
        Searches for a string in a nested JSON object (Python dict) starting from a specified parent,
        and returns the parent key under which the string resides.
        
        Parameters:
        - json_obj (dict): The JSON object to search through.
        - target_str (str): The string to look for.
        - start_parent (str): The parent key from where to start the search.
        - current_parent: The current parent key, used for recursion.
        
        Returns:
        - The parent key under which the string resides, or None if the string is not found.
        """
        if start_parent and current_parent is None:
            json_obj = json_obj.get(start_parent, {})
            
        for key, value in json_obj.items():
            if key == target_str:
                return current_parent
            if isinstance(value, dict):
                result = find_parent(value, target_str, None, key)
                if result:
                    return result
            elif isinstance(value, (list, set)):
                if target_str in value:
                    return key
        return None




    #########################################################################################################################################################################
    ############################################################################## GSR PARSING ##############################################################################
    #########################################################################################################################################################################
    #--> Gather Start/End Dates for the report

    # Read the CSV file to extract the start date and end date
    df = pd.read_csv(gsrFilePath, nrows=2)  # Reading only the first 2 rows

    # Extract the start date and end date from column C and D in row 2
    start_date_str = df.loc[1, 'Start Date']
    end_date_str = df.loc[1, 'End Date']

    start_date_dt = datetime.strptime(start_date_str, '%m/%d/%Y %I:%M %p')
    end_date_dt = datetime.strptime(end_date_str, '%m/%d/%Y %I:%M %p')
        
    # Create formatted date strings
    short_date_str = f"{start_date_dt.strftime('%Y-%m-%d')} to {end_date_dt.strftime('%Y-%m-%d')}"
    complete_date_str = f"{start_date_str} to {end_date_str}"
        
    #--> Prepare dictionaries for data entry
    df = pd.read_csv(gsrFilePath)
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']

        if site not in blacklisted_sites:
            if site not in MONTHLY_STATS:
                MONTHLY_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}}
            if site not in COMBINED_STATS:
                COMBINED_STATS[site] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            if Corporation_Totals_Name not in MONTHLY_STATS:
                MONTHLY_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0, 'Estimated Member Count': 0}}
            if Corporation_Totals_Name not in COMBINED_STATS:
                COMBINED_STATS[Corporation_Totals_Name] = {"Express Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Clean Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "Protect Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}, "UShine Wash":{'Count': 0, 'Price': 0, 'Quantity': 0, 'Amount': 0}}
            
            if site not in MONTHLY_SALES:
                MONTHLY_SALES[site] = {"NET Sales": 0}
            if site not in COMBINED_SALES:
                COMBINED_SALES[site] = {"Discounts": 0, "NET Sales": 0}
            if Corporation_Totals_Name not in MONTHLY_SALES:
                MONTHLY_SALES[Corporation_Totals_Name] = {"NET Sales": 0}
            if Corporation_Totals_Name not in COMBINED_SALES:
                COMBINED_SALES[Corporation_Totals_Name] = {"Discounts": 0, "NET Sales": 0}

    #--> Gather TOTAL sales & discounts details
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items:

            COMBINED_SALES[site]['NET Sales'] += amount if amount != None else 0
            COMBINED_SALES[Corporation_Totals_Name]['NET Sales'] += amount if amount != None else 0
            if category in Discount_Categories:
                COMBINED_SALES[site]['Discounts'] += amount if amount != None else 0
                COMBINED_SALES[Corporation_Totals_Name]['Discounts'] += amount if amount != None else 0

            # Handle the Query Server
            if site == 'Query Server':
                if find_parent(Website_PKG_Names, item_name, start_parent="Monthly") != None or category in Monthly_Total_Categories:
                    MONTHLY_SALES['Query Server']['NET Sales'] += amount if amount != None else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if amount != None else 0
            # Any other locations
            else: 
                if category in Monthly_Total_Categories:
                    MONTHLY_SALES[site]['NET Sales'] += amount if amount != None else 0
                    MONTHLY_SALES[Corporation_Totals_Name]['NET Sales'] += amount if amount != None else 0

            
                
    print(COMBINED_SALES)
    print(MONTHLY_SALES)

    #--> Gather Sales Mix
    """
    Calculating Combined, Monthly, and Retail Stats:

    Components:
    - ARM Redeemer: A package leader item triggered when a FastPass Plan customer pays.
    - ARM Plan Packages: Package leader items contained within the ARM Redeemer.
    - Wash Package: Actual wash service (e.g., 'Express', 'Protect').
    - ARM Redeemer Discount: Makes the Wash Package cost $0 for FastPass Plan customers.

    Transaction Flow:
    1. FastPass Plan customer arrives, triggering the ARM Redeemer.
    2. The ARM Redeemer contains ARM Plan Packages, each of which includes:
        1. A Wash Package (e.g., 'Express', 'Protect').
        2. An ARM Redeemer Discount, setting the Wash Package cost to $0.

    Stat Calculations:
    - Combined Stats (Retail + Monthly): Both retail and monthly transactions are recorded under 'Basic Washes' because they both trigger a Wash Package.
    - Monthly Stats: Tracked when an ARM Redeemer is activated (exclusive to monthly FastPass Plan customers).
    - Retail Stats: Obtained by subtracting Monthly Stats from Combined Stats.

    Formula:
    Retail Stats = Combined Stats - Monthly Stats

    Template:
    The template is Excel based and therefore can be set up early to calculate the retail data based on the input from this script for combined and monthly.
    """
    #NOTE: RETAIL statistics are determined based off of Combined & Monthly. Spreadsheet formulas will determine the Quantity and Totals as well.
    df = pd.read_csv(gsrFilePath)

    MonthlyAmounts = {}
    CombinedAmounts = {}

    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        item_name = row['Item Name']
        amount = row['Amount']
        count = row['Count']
        price = row['Price']
        quantity = row['Quantity']
        
        if site not in blacklisted_sites and item_name is not None and category not in blacklisted_items:

            washPkg = find_parent(ARM_Recharge_Names, item_name) # Member Count = ARM Recharges
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0
            washPkg = find_parent(Website_PKG_Names, item_name, "Monthly") # Member Count = ARM Recharges + Website Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0
            washPkg = find_parent(ARM_Sold_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] += quantity if quantity is not None else 0
            washPkg = find_parent(ARM_Termination_Names, item_name) # Member Count = ARM Recharges + Website Sld + ARM Sld - ARM Terminations
            if washPkg != None:
                MONTHLY_STATS[site][washPkg]['Estimated Member Count'] -= quantity if quantity is not None else 0

                MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Estimated Member Count'] -= quantity if quantity is not None else 0

            # Handle the Query Server
            if site == 'Query Server':
                # We're only going to show the amount of passes sold rather than the redemption for the
                # Query Server because you cannot redeem passes online.
                washPkg = find_parent(Website_PKG_Names, item_name, "Monthly")
                if washPkg != None:
                    MONTHLY_STATS['Query Server'][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Count'] += count if count is not None else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Amount'] += amount if amount is not None else 0
                    MONTHLY_STATS['Query Server'][washPkg]['Price'] += price if price is not None else 0

                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if count is not None else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if amount is not None else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if price is not None else 0)
                
                washPkg = find_parent(Website_PKG_Names, item_name, "Retail")
                if washPkg != None:
                    COMBINED_STATS['Query Server'][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    COMBINED_STATS['Query Server'][washPkg]['Count'] += count if count is not None else 0
                    COMBINED_STATS['Query Server'][washPkg]['Amount'] += amount if amount is not None else 0
                    COMBINED_STATS['Query Server'][washPkg]['Price'] += price if price is not None else 0

                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if count is not None else 0
                    COMBINED_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if amount is not None else 0
                    if not washPkg in MonthlyAmounts:
                        CombinedAmounts[washPkg] = []
                    CombinedAmounts[washPkg].append(price if price is not None else 0)

            # Any other locations
            else: 
                washPkg = find_parent(ARM_PKG_Names, item_name)
                # Calculate MONTHLY Stats using the redemption items
                if category in ["ARM Plans Redeemed", "Club Plans Redeemed"] and washPkg is not None:
                    print(item_name, washPkg)
                    MONTHLY_STATS[site][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    MONTHLY_STATS[site][washPkg]['Count'] += count if count is not None else 0
                    MONTHLY_STATS[site][washPkg]['Amount'] += amount if amount is not None else 0
                    MONTHLY_STATS[site][washPkg]['Price'] += abs(price) if price is not None else 0

                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Quantity'] += quantity if quantity is not None else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Count'] += count if count is not None else 0
                    MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Amount'] += amount if amount is not None else 0
                    if not washPkg in MonthlyAmounts:
                        MonthlyAmounts[washPkg] = []
                    MonthlyAmounts[washPkg].append(price if price is not None else 0) 
                elif category in ['Basic Washes'] and item_name is not None and item_name in COMBINED_STATS[site]:
                    #print("Quantitiy: ", quantity, COMBINED_STATS[site][item_name]['Quantity'])
                    COMBINED_STATS[site][item_name]['Quantity'] += quantity if quantity is not None else 0
                    #print(COMBINED_STATS[site][item_name]['Quantity'])
                    COMBINED_STATS[site][item_name]['Count'] += count if count is not None else 0
                    COMBINED_STATS[site][item_name]['Amount'] += amount if amount is not None else 0
                    COMBINED_STATS[site][item_name]['Price'] += abs(price) if price is not None else 0

                    print("Quantitiy: ", item_name,quantity, COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'])
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'] += quantity if quantity is not None else 0
                    print('Quan in JSON:', COMBINED_STATS[Corporation_Totals_Name][item_name]['Quantity'])
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Count'] += count if count is not None else 0
                    COMBINED_STATS[Corporation_Totals_Name][item_name]['Amount'] += amount if amount is not None else 0
                    if not washPkg in CombinedAmounts:
                        CombinedAmounts[item_name] = []
                    CombinedAmounts[item_name].append(price if price is not None else 0)
    print(MonthlyAmounts)
    for washPkg in ARM_Sold_Names:
        COMBINED_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(CombinedAmounts[washPkg])/len(CombinedAmounts[washPkg])
        MONTHLY_STATS[Corporation_Totals_Name][washPkg]['Price'] = sum(MonthlyAmounts[washPkg])/len(MonthlyAmounts[washPkg])

    
    #########################################################################################################################################################################
    ############################################################################## T&C PARSING ##############################################################################
    #########################################################################################################################################################################
    '''
    ChurnRates = {}
    df = pd.read_csv(trendsFilePath)
    for _, row in df.iterrows():
        date = row['Date']
        plan = row['Plan']
        site = row['Site']
        washes_per_day = row['Washes per day']
        upsells_dollars = row['Upsells (Dollars)']
        upsells = row['Upsells']
        plans_sold_dollars = row['Plans Sold (Dollars)']
        plans_sold = row['Plans Sold']
        recharges_dollars = row['Recharges (Dollars)']
        recharges = row['Recharges']
        plans_transferred_in = row['Plans Transferred In']
        plans_transferred_out = row['Plans Transferred Out']
        plans_transferred_in_dollars = row['Plans Transferred In (Dollars)']
        plans_transferred_out_dollars = row['Plans Transferred Out (Dollars)']
        suspended = row['Suspended']
        suspended_dollars = row['Suspended (Dollars)']
        resumed = row['Resumed']
        resumed_dollars = row['Resumed (Dollars)']
        plans_terminated = row['Plans Terminated']
        plans_terminated_dollars = row['Plans Terminated (Dollars)']
        plans_discontinued = row['Plans Discontinued']
        cc_declining_expired = row['CC Declining + CC Expired']
        expired_members = row['Expired Members']
        members_per_day = row['Members per day']
        revenue = row['Revenue']
        pass_wash_percentage = row['Pass Wash Percentage']
        churn_rate = row['Churn Rate']

    Need a formula to determine the churn rate for the corporatiion based off of the churn rate from the individual stores. 

    '''
    #########################################################################################################################################################################
    ############################################################################## XCL LOADING ##############################################################################
    #########################################################################################################################################################################
    #--> Utility Functions to copy a range of cells from one location to another, including their styles and formulas
    def adjust_cell_reference(match, row_shift, col_shift):
        cell_ref = match.group(0)
        col_ref, row_ref = re.match(r"([A-Z]+)([0-9]+)", cell_ref).groups()
        new_col = get_column_letter(column_index_from_string(col_ref) + col_shift)
        new_row = str(int(row_ref) + row_shift)
        return new_col + new_row

    def adjust_formula(formula, row_shift, col_shift):
        pattern = r"(\$?[A-Z]+\$?[0-9]+)"
        return re.sub(pattern, lambda match: adjust_cell_reference(match, row_shift, col_shift), formula)

    def copy_cells(ws, start_row, end_row, start_col, end_col, target_start_row, target_start_col):
        row_shift = target_start_row - start_row
        col_shift = target_start_col - start_col

        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)):
            for j, cell in enumerate(row):
                target_cell = ws.cell(row=target_start_row + i, column=target_start_col + j)
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Adjust the formula
                    adjusted_formula = adjust_formula(cell.value[1:], row_shift, col_shift)
                    target_cell.value = f"={adjusted_formula}"
                else:
                    target_cell.value = cell.value

                # Copy cell style if it has one
                if cell.has_style:
                    target_cell._style = copy(cell._style)
                
                # Copy font, border, fill, number format, etc.
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)


    #--> Populate Worksheet Function to fill in each location block
    def populate_salesmix_worksheet(worksheet, start_row, start_col, location):

        #--> Assign location name as the new block name
        if location != 'Query Server':
            worksheet.cell(row=start_row, column=start_col, value=location)
        else:
            worksheet.cell(row=start_row, column=start_col, value="E-Commerce Website")
        
        #--> BLOCK: Retail Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            worksheet.cell(row=start_row+3+i, column=start_col+3, value=pkgProps['Price'])
        
        #--> BLOCK: Monthly Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(MONTHLY_STATS[location].items()):
            worksheet.cell(row=start_row+12+i, column=start_col+1, value=pkgProps['Quantity'])
            worksheet.cell(row=start_row+12+i, column=start_col+5, value=pkgProps['Estimated Member Count'])
        worksheet.cell(row=start_row+17, column=start_col+1, value=MONTHLY_SALES[location]['NET Sales'])

        #--> BLOCK: Combined Stats
        #### Need to fill in the prices (average price) because the Quantity will be determined by the excel formulas.
        for i, (washPkg, pkgProps) in enumerate(COMBINED_STATS[location].items()):
            if location != 'Query Server':
                worksheet.cell(row=start_row+21+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+21+i, column=start_col+3, value=pkgProps['Price'])
            else:
                worksheet.cell(row=start_row+3+i, column=start_col+1, value=pkgProps['Quantity'])
                worksheet.cell(row=start_row+3+i, column=start_col+3, value=pkgProps['Price'])
                worksheet.cell(row=start_row+21+i, column=start_col+1, value=pkgProps['Quantity']+MONTHLY_STATS['Query Server'][washPkg]['Quantity'])
                worksheet.cell(row=start_row+21+i, column=start_col+3, value=pkgProps['Price'])
        
        worksheet.cell(row=start_row+26, column=start_col+1, value=COMBINED_SALES[location]['NET Sales'])
        worksheet.cell(row=start_row+26, column=start_col+3, value=COMBINED_SALES[location]['Discounts'])
        
        #--> Query Server Modifications:
        if location == "Query Server":
            worksheet.cell(row=start_row+11, column=start_col+1, value="New Passes Sold")
            worksheet.cell(row=start_row+16, column=start_col, value="Total ARM Plans Sold")
            worksheet.cell(row=start_row+20, column=start_col+1, value="Items Sold")
            worksheet.cell(row=start_row+25, column=start_col, value="Total Items Sold for Period")
        
        


    #--> Modify the template and populate all blocks with site data
    wb = openpyxl.load_workbook(templateFilePath)
    ws = wb['Sales Mix by Location']

    start_col = 1
    ws.cell(row=29, column=2, value=complete_date_str)
    locations = list(MONTHLY_STATS.keys())
    locations.remove("Corporation Totals")
    locations.remove("Query Server")
    for location in locations:
        copy_cells(ws, 1, 27, 1, 6, 1, start_col)
        populate_salesmix_worksheet(ws, 1, start_col, location)
        start_col += 7
    
    copy_cells(ws, 1, 27, 1, 7, 31, 1)
    populate_salesmix_worksheet(ws, 31, 1, Corporation_Totals_Name)
    copy_cells(ws, 1, 27, 1, 7, 31, 8)
    populate_salesmix_worksheet(ws, 31, 8, "Query Server")

    
        
    #--> Discounts Worksheet Stuff
    # Define styles
    int_style = NamedStyle(name='integer_style', number_format='#,##0')
    accounting_style = NamedStyle(
        name='accounting_style', 
        number_format='#,##0.00',  # or you could use 'Accounting'
        alignment=Alignment(horizontal='right')
    )
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    # Modify the populate_discountmix_worksheet function to include formatting
    def populate_discountmix_worksheet(worksheet, start_row, start_col, location, discount_data):
        worksheet.cell(row=start_row, column=start_col, value=location)
        
        # Adding headers
        headers = ["Discount Name", "Quantity Used", "Price*", "Total Discount Dollars", "Discount Distribution"]
        for j, header in enumerate(headers):
            worksheet.cell(row=start_row + 2, column=start_col + j, value=header)
            
        total_quantity = 0
        total_amount = 0

        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            total_quantity += discount_info['Quantity'] if discount_info['Quantity'] is not None else 0
            total_amount += discount_info['Amount'] if discount_info['Amount'] is not None else 0
        
        for i, (discount_name, discount_info) in enumerate(discount_data.items()):
            if discount_name is not None:

                # Add discount name
                worksheet.cell(row=start_row+3+i, column=start_col, value=discount_name)
                
                # Add and format Quantity
                cell_quantity = worksheet.cell(row=start_row+3+i, column=start_col+1, value=discount_info['Quantity'])
                cell_quantity.style = int_style
                
                # Add and format Price
                cell_price = worksheet.cell(row=start_row+3+i, column=start_col+2, value=discount_info['Price'])
                cell_price.style = accounting_style
                
                # Add and format Amount
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+3, value=discount_info['Amount'])
                cell_amount.style = accounting_style

                # Add and format Distribution
                print(discount_info['Quantity'], total_quantity, discount_name)
                distribution_value = discount_info['Quantity'] / total_quantity if total_quantity != 0 else 0
                cell_amount = worksheet.cell(row=start_row+3+i, column=start_col+4, value=distribution_value)
                cell_amount.style = percent_style  # Apply the percentage style
                
                

        # Add and format "Total"
        total_label_cell = worksheet.cell(row=start_row+4+i, column=start_col, value="Total")
        total_label_cell.font = Font(bold=True)

        # Add and format Total Quantity
        cell_total_quantity = worksheet.cell(row=start_row+4+i, column=start_col+1, value=total_quantity)
        cell_total_quantity.style = int_style  # If you have defined int_style elsewhere
        cell_total_quantity.font = Font(bold=True)

        # Add and format Total Amount
        cell_total_amount = worksheet.cell(row=start_row+4+i, column=start_col+3, value="${:.2f}".format(total_amount))
        cell_total_amount.style = accounting_style  # If you have defined accounting_style elsewhere
        cell_total_amount.font = Font(bold=True)

        # Add a Price disclaimer explaining that some prices are averages
        if location == Corporation_Totals_Name:
            disclaim1 = worksheet.cell(row=start_row+4+i+2, column=start_col, value="NOTES")
            disclaim2 = worksheet.cell(row=start_row+4+i+3, column=start_col, value="Price*: Some 'Open Price' items such as 'Misc. Wash Discount' show the average amount that was paid out over the period under that discount item.")
            disclaim3 = worksheet.cell(row=start_row+4+i+4, column=start_col, value="Discounts that are not applied (are missing on the GSR for the location while parsing; thus quantity is 0) will not be shown on the report for that location.")
            disclaim4 = worksheet.cell(row=start_row+4+i+5, column=start_col, value="")
            disclaim1.font = Font(bold=True)
            disclaim2.font = Font(bold=True)
            disclaim3.font = Font(bold=True)
            disclaim4.font = Font(bold=True)

        tableName = re.sub(r'\W+', '', location.replace(" ", ""))
        # Adding table and table banding
        tab = Table(displayName=tableName, ref=f"{worksheet.cell(row=start_row + 2, column=start_col).coordinate}:{worksheet.cell(row=start_row + 4 + i, column=start_col + 4).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium23", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    # Read the input CSV to a DataFrame
    df = pd.read_csv(gsrFilePath)

    # Aggregate the data
    discount_data = {}
    miscDisctAverages = {}
    discountTotals = {}
    discount_data[Corporation_Totals_Name] = {}
    for _, row in df.iterrows():
        site = row['Site']
        category = row['Report Category']
        discount_name = row['Item Name']
        quantity = row['Quantity']
        price = row['Price']
        amount = row['Amount']

        if (category in Discount_Categories and 
            site not in blacklisted_sites and 
            pd.notna(discount_name) and 
            pd.notna(quantity) and 
            pd.notna(price) and 
            pd.notna(amount)):
            
            if site not in discount_data:
                discount_data[site] = {}
                
            if discount_name not in discount_data[site]:
                discount_data[site][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            if discount_name not in discount_data[Corporation_Totals_Name]:
                discount_data[Corporation_Totals_Name][discount_name] = {'Quantity': 0, 'Price': 0, 'Amount': 0}
            
            discount_data[site][discount_name]['Quantity'] += quantity
            discount_data[Corporation_Totals_Name][discount_name]['Quantity'] += quantity
            if discount_name not in miscDisctAverages:
                miscDisctAverages[discount_name] = []
            miscDisctAverages[discount_name].append(price if pd.notna(price) else 0)
            discount_data[site][discount_name]['Amount'] += amount
            discount_data[Corporation_Totals_Name][discount_name]['Amount'] += amount
    
    # Average the prices
    # Used for Misc. discounts to see the average amount of money paid out to customers
    locations = list(discount_data.keys())

    for location in locations:
        for discount in discount_data[location]:
            try:
                discount_data[location][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
                discount_data[Corporation_Totals_Name][discount]['Price'] = sum(miscDisctAverages[discount])/len(miscDisctAverages[discount])
            except:
                discount_data[location][discount]['Price'] = 0

    # Load the template workbook
    ws = wb['Discount Mix by Location']
    print(discount_data)
    # Initial starting column for the first location
    start_col = 1

    # Populate the worksheet
    for location, data in discount_data.items():
        copy_cells(ws, 1, 3, 1, 5, 1, start_col)
        populate_discountmix_worksheet(ws, 1, start_col, location, data)
        start_col += 6
    
    wbName = f"Sales Mix - {short_date_str}.xlsx"
    return wb, wbName

wb, wbName = createSalesMixSheet("input.csv", "input_tac.csv", "Sales_Mix_Template.xlsx")
wb.save(wbName)
